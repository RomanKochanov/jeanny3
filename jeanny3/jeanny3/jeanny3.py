# -*- coding: utf-8 -*-

import os
import re
import sys
import csv      
import json
import copy
import shutil
import string
import hashlib
import uuid as uuidmod
import itertools as it
import functools as ft

from itertools import cycle
from datetime import date, datetime
from warnings import warn,simplefilter

# increase the csv field size limit
#csv.field_size_limit(sys.maxsize) # throws error on some versions of Python
csv.field_size_limit(1000000000)

simplefilter('always', UserWarning) # make warnings display always

# Adding a JSON-serialization to the Collection type
from json import JSONEncoder
def _default(self, obj):
    return getattr(obj.__class__, "export_to_json", _default.default)(obj)
_default.default = JSONEncoder().default
JSONEncoder.default = _default

# Simple dictionaty-based DBMS for the data-mining.
# This DBMS inherits a shema-free database ideology (like MongoDB for instance).
# Each document in the collection has it's own unique ID number.
       
# Structure of a collection:
#
# Document = {'__ID__':Integer, '__DATA__':Dictionary} : Dictionary
# Collection = [Document1, Document2, ...]

# Filtering:
# >> Col = Collection(Col1)
# >> IDs = Col.IDs(filter='var['a']+var['b']>10') # get IDs with condition 
# >> Data = Col.get(IDs)
# This way of defining filters must be avoided in the release because of 
# the obvious vulnerabilities!

# VERSION 3.0

#from collections import OrderedDict # doesn't help a lot
#from addict import Dict # takes less memory than OrderedDict

# !!! write more efficient engine on the top of built-in sqlite module?
# https://pypi.python.org/pypi/sqlite-schemaless/0.1.2
# http://yserial.sourceforge.net/
# Good idea, but what to do about open data structure? We need a version control 
# layer on the top of the database engine!

__version__ = '3.0'
print('jeanny, Ver.'+__version__)

FILENAME_ID = '$FILENAME$' # parameter defining the filename
ITEM_ID = '$UUID$' # id that uniquely identifies each item (used in redistribution of items between collections)

SETTINGS = {
    'FILENAME_ID': FILENAME_ID,
    'ITEM_ID': ITEM_ID,
    'DEBUG': False,
    #'PLOTTING_BACKEND': 'Agg',
}

# stub for Python 3: redefining unicode
try:
    unicode
except NameError:
    unicode = str # in Python 3 str is unicode

def uuid():
    # http://stackoverflow.com/questions/2759644/python-multiprocessing-doesnt-play-nicely-with-uuid-uuid4    
    return str(uuidmod.UUID(bytes=os.urandom(16), version=4))    
    
# https://stackoverflow.com/questions/11875770/how-to-overcome-datetime-datetime-not-json-serializable-in-python
def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    if isinstance(obj, (datetime, date)):
        serial = obj.isoformat()
        return serial
    raise TypeError ("Type %s not serializable" % type(obj))

def is_identifier(token):    
    return re.match('^[a-zA-Z_][\w_]*$',token)
    
def process_exp(raw_expression): # think about more proper name
    # change this ('a**2+2') to this ('var["a"]**2+2')
    import shlex
    tokens = list(shlex.shlex(raw_expression))    
    for i,token in enumerate(tokens):
        if is_identifier(token): tokens[i] = 'var["%s"]'%token
    return ''.join(tokens)

#class TabObject:
#    """
#    Class for string representation.
#    Used for the output of tabulate function.
#    """
#    def __init__(self,st):
#        self.__string__ = st
#    def tostr(self):
#        return self.__string__
        
class Collection:
    
    #ID_REP = '$id'

    def __init__(self,path=None,**argv):
        self.initialize(path=path,**argv)
        
    def __getitem__(self,ID):
        return self.getitem(ID)
        
    def __iter__(self):
        return iter(self.getitems())
        
    def __len__(self):
        return len(self.__dicthash__)

    #def export_to_json(self): # old and useless version
    #    """
    #    Adding a serialization to the Collection type (see above).
    #    """
    #    if self.__path__ in {'',None}:
    #        warn('The collection is not actually saved')
    #    return {
    #        '__class__':'Collection',
    #        '__name__':self.__name__,
    #        '__type__':self.__type__,
    #        '__path__':self.__path__}

    def export_to_json(self): # old and useless version
        """
        Adding a serialization to the Collection type (see above).
        """
        return self.__dicthash__
                
    def clear(self): # synonym for __init__ (TODO: get rid of it)
        self.initialize() 
        # XXX: use custom method instead of __init__
        # in order not to cause conflicts with inherited classes
        
    def __repr__(self):
        
        #buf = '%s(%s): %s\n'%(self.__name__,self.__type__,self.__path__)
        #buf += json.dumps(self.keys(),indent=2)
        #print(buf)
        #return buf
        #print(self.tabulate())
        #return self.tabulate(raw=True)
        
        #ids = self.ids()
        #buf = self.tabulate(raw=True,IDs=ids[:20])
        #if len(ids)>20: buf+='\n...'
        #return buf
        
        return 'Collection (%d lines)'%len(self.ids())

    def initialize(self,path=None,fmt=None,name='Default',**argv):
        self.maxid = -1
        self.order = [] # order of columns (optional)
        self.types = None # numpy-compatible typing header (for export to DB)
        if path is not None:
            if type(path) is str:
                if not fmt: raise Exception('Collection type/format is not specified')
                if fmt=='csv':
                    self.import_csv(path,**argv)
                elif fmt=='folder':                
                    self.import_folder(path,**argv)
                elif fmt=='jsonlist':
                    self.import_json_list(path,**argv)
                elif fmt=='xlsx':
                    self.import_xlsx(path,**argv)
                elif fmt=='fixcol':
                    self.import_fixcol(path,**argv)
                else:
                    raise Exception('Unknown type: %s'%fmt)
                self.__type__ = fmt
                self.__path__ = path
                self.__name__ = name
            elif type(path) is list:
                self.__dicthash__ = {}
                self.__name__ = ''
                self.__path__ = './'
                self.__type__ = '__init__'
                self.update(path)
                if path: self.order = list(path[0].keys())
            else:
                raise Exception('unknown content type: %s'%type(path))
        else:
            self.__dicthash__ = {}
            self.__name__ = ''
            self.__path__ = './'
            self.__type__ = '__init__'
            #self.__dicthash__ = OrderedDict()
            #self.__dicthash__ = Dict() # !!! ATTENTION !!! when fetching non-existing item, creates it => non-standard behaviour
            
    def export(self,path=None,type=None,**argv):
        """
        Export the collection using several available strategies.
        """
        if path is None: path = self.__path__
        if type is None: type = self.__type__
        if type=='csv':
            self.export_csv(path,**argv)
        elif type=='folder':                
            self.export_folder(path,**argv)
        elif type=='jsonlist':
            self.export_json_list(path,**argv)
        elif type=='xlsx':
            self.export_xlsx(path,**argv)
        else:
            raise Exception('Unknown type: %s'%type)        
        
    #def __deepcopy__(self):
    #    """
    #    Function override for the deep copy.
    #    https://stackoverflow.com/questions/1500718/what-is-the-right-way-to-override-the-copy-deepcopy-operations-on-an-object-in-p
    #    """
    #    pass
    
    def copy(self,name='Default',path=''): # CAN BE BUGGY!!
        """
        Copy the collection using the deep copy feature.
        https://www.geeksforgeeks.org/copy-python-deep-copy-shallow-copy/
        """
        col = copy.deepcopy(self)
        col.__name__ = name
        col.__path__ = path # don't use the parent's path by default
        return col

    def setorder(self,order):
        self.order = order
        
    def setfloatfmt(self,floatfmt):
        self.floatfmt = floatfmt
            
    def getitem(self,ID):
        #if type(ID) in [list,tuple]:
        #    if len(ID)>1:
        #        raise Exception('ID must be aither scalar or list of 1 element')
        #    ID = ID[0]
        if ID not in self.__dicthash__:
            #raise Exception('no such ID in __dicthash__: %s'%ID)
            raise KeyError('no such ID in __dicthash__: %s'%ID) # I think that this will mess up the workflow
            #return None
        return self.__dicthash__[ID]
    
    def getitems(self,IDs=-1,mode='strict'):
        """
        Empty IDs must lead to the empty item list!
        Modes: strict,greedy,silent
        """
        if IDs == -1:
            IDs = self.ids()
        buffer = []
        for ID in IDs:
            # some dictionaries create item when it is not found,
            # so do explicit check on item's existence
            if ID not in self.__dicthash__:
                if mode=='strict':
                    raise Exception('no such ID in __dicthash__: %s'%ID)
                elif mode=='silent':
                    continue
                elif mode=='greedy':
                    buffer.append(None)
            else:
                buffer.append(self.__dicthash__[ID])
        return buffer
            
    def getfreeids(self,n):
        # Generate n IDs which don't exist in the collection.
        # TODO: optimize
        idmin = self.maxid + 1
        self.maxid += n # REMOVE THIS FROM HERE
        return list(range(idmin,idmin+n)) # gives error in Python 3 without the list() wrapper
        
    def shuffle(self,n,IDs=-1): # CHANGE NAME!!!
        """
        Shuffle the collection using the round-robin strategy.
        """
        if IDs==-1:
            IDs = self.ids()
        lst_indexes = range(len(IDs))
        shuffled_list = []
        for i in range(n):
            subindex = range(i,len(IDs),n)
            shuffled_sublist = [IDs[i] for i in subindex]
            shuffled_list.append(self.subset(shuffled_sublist))
        return shuffled_list

    def ids(self,filter='True',proc=False):
        if proc: # allows filter be much more simple to input
            filter = process_exp(filter) # experimental
        if type(filter)==str: # simple
            expr = eval('lambda var: ' + filter)
        else:
            expr = filter # advanced
        # this nice trick is taken from stack overflow:
        # http://stackoverflow.com/questions/12467570/python-way-to-speed-up-a-repeatedly-executed-eval-statement?lq=1
        id_list = []
        for ID in self.__dicthash__:
            var = self.__dicthash__[ID]
            #try: # risky
            #    flag = expr(var)
            #except KeyError:
            #    flag = False
            flag = expr(var)
            if flag:
                id_list.append(ID)
        return id_list
        
#    def keys(self):
#        # old version
#        keys = set()
#        for ID in self.__dicthash__:
#            keys = keys.union(self.__dicthash__[ID].keys())
#        keys = list(keys); keys.sort()
#        keys = tuple(keys)  
#        return keys
    
    def keys(self):
        # new version, slow but more informative
        keys = {}
        for ID in self.__dicthash__:
            for key in self.__dicthash__[ID].keys():
                if key not in keys:
                    keys[key] = 1
                else:
                    keys[key] += 1
        return keys

#    def keys(self):
#        # new version #2, slightly faster
#        keys_ = {}
#        # map
#        for ID in self.__dicthash__:
#            k = tuple(self.__dicthash__[ID].keys())
#            if k not in keys_:
#                keys_[k] = 1
#            else:
#                keys_[k] += 1
#        # reduce
#        keys = {}
#        for key_tuple in keys_:
#            for key in key_tuple:
#                if key not in keys:
#                    keys[key] = keys_[key_tuple]
#                else:
#                    keys[key] += keys_[key_tuple]
#        return keys

    def subset(self,expr=None): # keys must be the same as in the original collection
        if type(expr) is type(lambda:None):
            IDs = self.ids(expr)
        elif type(expr) in {list,tuple,set,dict}:
            IDs = list(expr)
        elif expr is None:
            IDs = self.ids()
        else:
            raise Exception('unknown type of input expression')
        #new_coll = Collection()
        #items = [self.__dicthash__[ID] for ID in IDs]
        #new_coll.update(items,IDs)
        #new_coll.order = self.order
        #return new_coll
        new_coll = Collection()
        new_coll.__dicthash__ = {ID:self.__dicthash__[ID] for ID in IDs}
        #new_coll.__dicthash__ # ??
        new_coll.order = self.order
        new_coll.maxid = max(IDs) if len(IDs)!=0 else -1
        return new_coll
        
    def subset_group(self,grpi):
        """
        Create subset using the group index.
        In this case the subset IDs will be the concatenation
        of IDs of group keys.
        The resulting collection is a "proxy", wich means that
        it's items are the same that the items of self.
        """
        new_coll = Collection()
        dicthash = {}
        maxid = None
        for key in grpi:
            ids = grpi[key]
            for id_ in ids:
                dicthash[id_] = self.__dicthash__[id_]
                if maxid==None or id_>maxid: # TODO: get rid of maxid everywhere
                    maxid = id_
        new_coll.__dicthash__ = dicthash
        new_coll.order = self.order
        new_coll.maxid = maxid
        return new_coll
        
    def slice(self,colnames=None):
        """
        Produce new collection contaning only specified colnames.
        If colnames are not supplied, then a full copy of collection
        is returned.
        """
        if type(colnames) is str:
            colnames = [colnames]
        col = Collection()
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            if colnames:
                item_ = {k:item[k] for k in colnames if k in item}
            else:
                item_ = {k:item[k] for k in item}
            col.__dicthash__[ID] = item_ # item needs to be added
            #if item_:
            #    col.__dicthash__[ID] = item_ # commenting out this will allow empty items!!! (needs attention)
        if colnames:
            col.order = colnames
        else:
            col.order = self.order.copy()
        return col
        
    def map(self,expr={}):
        """
        Map collection items to different keys and return a new collection.
        The expression "expr" should be a either a static mapping dictionary,
        or a lambda function taking a list of keys and returning a mapping dictionary.
        Example 1 (static):
            expr = {'a':'b','c':'d'}
        Example 2 (lambda):
            expr = lambda keys: ['_'+k for k in keys]
        Default mapping doesn't change key names at all.
        """
        if type(expr) is dict:
            expr_ = lambda keys: expr
        else:
            expr_ = expr
        col = Collection()
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            keys = list(item.keys())
            mapping = expr_(keys)
            keys_mapped = []
            keys_unmapped = []
            for k in keys:
                if k in mapping:
                    keys_mapped.append(k)
                else:
                    keys_unmapped.append(k)
            item_ = {}
            for k in keys_unmapped:
                item_[k] = item[k]
            for k in keys_mapped:
                item_[mapping[k]] = item[k]            
            if item_:
                col.__dicthash__[ID] = item_
        # handle order
        mapping = expr_(self.order)
        keys_mapped = []
        keys_unmapped = []
        for k in self.order:
            if k in mapping:
                keys_mapped.append(mapping[k])
            else:
                keys_unmapped.append(k)
        col.order = keys_unmapped + keys_mapped
        return col
        
    def cast(self,type_dict,IDs=-1):
        if IDs==-1:
            IDs = self.ids()
        nchanged = 0
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            flag_changed = False
            for col in type_dict:
                if not col in var:
                    continue
                flag_changed = True
                tp = type_dict[col]
                if col in var:
                    var[col] = tp(var[col])
            if flag_changed:
                nchanged += 1
        return {'changed':nchanged}
    
    def batch_(self,expr,IDs=-1):
        if IDs==-1:
            IDs = self.ids()
        ##### OPTIMIZE!!!!!! expr is parsed on each iteration
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            exec(expr) # very slow!!!
    
    def assign(self,par,expr,IDs=-1):
        """
        Create new parameter and assign 
        some initial value that may depend on
        other parameters within the item.
        ATTENTION: WILL BE DEPRECATED
        """
        if IDs==-1:
            IDs = self.ids()
        if type(expr)==str: # simple
            expr = eval('lambda var: ' + expr)
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            var[par] = expr(var)
        #self.__order__.append(par)
        if par not in self.order:
            self.order.append(par)
            
    def assign_(self,expr,IDs=-1): 
        """
        More flexible and convenient version of assign.
        Takes a lambda, function or callable object 
        as an input, and returns a dictionary
        of the type {'a1':val1,'a2':val2,...}
        where a1,a2,... are new/existing fields, and 
        val1, ... are values to be assigned to those fields.
        The expression takes current item as an input.
        The fields may vary depending on the current item.
        """
        if IDs==-1:
            IDs = self.ids()
        if type(expr)==str: # simple
            expr = eval('lambda var: ' + expr)
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]            
            vals = expr(var)
            for par in vals:
                var[par] = vals[par]
        #self.__order__ += list(dct.keys())

    def assign__(self,dct,IDs=-1):  # expr => dct
        """
        More flexible and convenient version of assign.
        New version takes a dictionary of functions 
        as an input, and returns a dictionary
        of the type {'a1':val1,'a2':val2,...}
        where a1,a2,... are new/existing fields, and 
        val1, ... are values to be assigned to those fields.
        The expression takes current item as an input.
        The fields may vary depending on the current item.
        """
        if IDs==-1:
            IDs = self.ids()
        #if type(expr)==str: # simple
        #    expr = eval('lambda var: ' + expr)
        if type(dct) is not dict:
            raise Exception('dictionary is expected at input')
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]            
            #vals = expr(var)
            #for par in vals:
                #var[par] = vals[par]
            for par in dct:
                var[par] = dct[par](var)
        #self.__order__ += list(dct.keys())
        
    def index(self,expr): # ex-"reform"
        """
        !!! ATTENTION !!!
        Reforms core dictionary by assigning another parameter as key.
        New key MUST be defined for all items in collection.
        Don't use this method you are not sure about consequences.
        Parameter expr can be a string or function. 
        __________________________________________
        !!! TODO: in next index implementation add 
        expressions instead of the field  name.
        __________________________________________
        """
        if type(expr) == str:
            new_id_func = eval('lambda var: var["%s"]'%expr)
            #new_id_func = eval('lambda var: %s'%expr) # this is fucking confusing, don't uncomment it anymore please
        elif type(expr) in {tuple,list}:
            new_id_func = lambda v: tuple([v[k] for k in expr])
        else:
            new_id_func = expr # user-supplied function on item
        # check uniqueness of a new index key
        new_id_vals = []
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            new_id_vals.append(new_id_func(item))
        if len(new_id_vals)!=len(set(new_id_vals)):
            raise Exception('new index is not unique')
        # if there is no duplicates, proceed further
        __dicthash__ = self.__dicthash__ # backup dict hash
        self.__dicthash__ = {}
        keys = list(__dicthash__.keys()) # Python 3 has special object dict_keys
        for ID in keys:
            item = __dicthash__.pop(ID)
            ID_ = new_id_func(item)
            #item['__id__'] = ID_
            self.__dicthash__[ID_] = item
        return self
        
    def get(self,ID,colname):
        """
        Get an element from an item with given ID.
        The point-reference is supported in colnames, i.e.
        "a.b" will search for parameter "b" in the object "a".
        Point-referencing can be multiple.
        """
        if ID not in self.__dicthash__:
            raise Exception('ID=%s is not in dicthash'%str(ID))        
        item = self.__dicthash__[ID]
        chain = colname.split('.')
        cur_obj = item[chain[0]]
        for e in chain[1:]:
            #cur_obj = eval('cur_obj.%s'%e) # atrocious (and slow)
            cur_obj = getattr(cur_obj,e) 
        return cur_obj
        
    def getcols(self,colnames,IDs=-1,strict=True,mode=None,
                functions=None,process=None): # get rid of "strict" argument in ver. 4.0
        """
        Extract columns from collection.
        If parameter "strict" set to true,
        no exception handling is performed.
        The "functions" parameter is a dictionary containing
        the functions on the item. It also should be present in "colnames".
        Another update: now colname can have a properties,
        such as "col.var"
        __ID__ is a special parameter which corresponds to the local __dicthash__ ID.
        """
        # mode options: 'strict', 'silent', 'greedy'   # add this to docstring in ver. 4.0.
        if not mode: mode = 'strict' if strict else 'silent' 
        if not functions: functions = {}
        #print('%s mode'%mode)
        if IDs==-1:
            IDs = self.ids()
        if type(colnames) is str:
            colnames = [colnames]
        elif type(colnames) is not list:
            raise Exception('Column names should be either list or string')
        cols = []
        for colname in colnames:
            if type(colname) not in [str,unicode]:
                raise Exception('Column name should be a string')
            cols.append([])
        for ID in IDs:
            for i,colname in enumerate(colnames):
                if colname == '__ID__':
                    cols[i].append(ID)
                    continue                    
                try:
                    #cols[i].append(self.__dicthash__[ID][colname]) % old
                    if colname not in functions:
                        cols[i].append(self.get(ID,colname)) # "get" this should be a method of item in the next version of Jeanny
                    else:
                        cols[i].append(functions[colname](self.__dicthash__[ID]))
                except (KeyError, AttributeError) as e: 
                    if mode=='strict':
                        raise e
                    elif mode=='silent':
                        pass
                    elif mode=='greedy':
                        cols[i].append(self.__dicthash__[ID].get(colname))
                    else:
                        raise Exception('unknown mode: %s'%mode)
                    #if strict: raise e # old version with "strict" argument
        if process:
            cols = [process(col) for col in cols]
        return cols
        
    def getcol(self,colname,IDs=-1,strict=True,mode='greedy',functions=None): # get rid of "strict" argument in ver. 4.0
        """
        Wrapper for a single-column call.
        """
        colnames = [colname,]
        return self.getcols(colnames=colnames,IDs=IDs,strict=strict,mode=mode,functions=functions)[0]
        
    def getrows(self,colnames,IDs=-1,strict=True,mode=None,
                functions=None,process=None): # get rid of "strict" argument in ver. 4.0
        """
        Extract rows from collection.
        If parameter "strict" set to true,
        no exception handling is performed.
        The "functions" parameter is a dictionary containing
        the functions on the item. It also should be present in "colnames".
        Another update: now colname can have a properties,
        such as "col.var"
        __ID__ is a special parameter which corresponds to the local __dicthash__ ID.
        """
        # mode options: 'strict', 'silent', 'greedy'   # add this to docstring in ver. 4.0.
        if not mode: mode = 'strict' if strict else 'silent' 
        if not functions: functions = {}
        #print('%s mode'%mode)
        if IDs==-1:
            IDs = self.ids()
        if type(colnames) is str:
            colnames = [colnames]
        elif type(colnames) is not list:
            raise Exception('Column names should be either list or string')
        for colname in colnames:
            if type(colname) not in [str,unicode]:
                raise Exception('Column name should be a string')
        rows = []
        for ID in IDs:
            row = []
            for i,colname in enumerate(colnames):
                if colname == '__ID__':
                    row.append(ID)
                    continue
                try:
                    if colname not in functions:
                        row.append(self.get(ID,colname)) # "get" this should be a method of item in the next version of Jeanny
                    else:
                        row.append(functions[colname](self.__dicthash__[ID]))
                except (KeyError, AttributeError) as e: 
                    if mode=='strict':
                        raise e
                    elif mode=='silent':
                        row.append(None)
                    elif mode=='greedy':
                        row.append(None)
                    else:
                        raise Exception('unknown mode: %s'%mode)
                    #if strict: raise e # old version with "strict" argument
            rows.append(tuple(row))
        if process:
            rows = [process(row) for row in rows]
        return rows
        
    def splitcol(self,colname,newcols=None):
        if newcols is None:
            vals = self.getitem(next(iter(self.__dicthash__.keys())))[colname]
            newcols = [colname+'_%d'%i for i,_ in enumerate(vals)]
        for i,cname in enumerate(self.order):
            if cname==colname: self.order.pop(i)
        self.order += newcols
        for item in self.getitems():
            vals = item[colname]
            for newcol,val in zip(newcols,vals):
                item[newcol] = val
            del item[colname]
                       
    def deletecol(self,colname):
        """
        Delete column from collecition.
        """        
        self.deletecols(colname)
            
    def deletecols(self,colnames):
        """
        Delete columns from collecition.
        """        
        if type(colnames) not in [list,tuple]:
            colnames = [colnames]
        for i,cname in enumerate(self.order):
            if cname in colnames: self.order.pop(i)
        for item in self.getitems():
            for colname in colnames:
                if colname in item:
                    del item[colname]
                    
    def renamecol(self,oldname,newname):
        """
        Rename column of collection.
        """
        # delete from items
        for item in self.getitems():
            if oldname in item:
                item[newname] = item.pop(oldname)
        # delete from order
        findall = lambda lst,val: [i for i,x in enumerate(lst) if x==val]
        if self.order is not None:
            indices = findall(self.order,oldname)
        for ind in indices:
            self.order[ind] = newname
        
    def split(self,colname,vals):
        """ Split collection using sharding of column by values.
            Column 'colname' must have numeric format. """
        cols = []
        for i,val in enumerate(vals):
            if i==0:
                col = self.subset(self.ids(lambda v:v[colname]<=val))
            else:
                val_ = vals[i-1]
                col = self.subset(self.ids(lambda v:val_<v[colname]<=val))
            cols.append(col)
        col = self.subset(self.ids(lambda v:v[colname]>val))
        cols.append(col)
        return cols

    def tabulate(self,colnames=None,IDs=-1,mode='greedy',fmt='simple',file=None,functions=None,raw=False,floatfmt=None): # switched default to "greedy" instead of "strict"
        """
        Supported table formats are:
        
        - "plain"
        - "simple"
        - "grid"
        - "fancy_grid"
        - "pipe"
        - "orgtbl"
        - "jira"
        - "psql"
        - "rst"
        - "mediawiki"
        - "moinmoin"
        - "html"
        - "latex"
        - "latex_booktabs"
        - "textile"

        More info on usage of Tabulate can be found at 
        https://pypi.org/project/tabulate/
        """
                
        try:
            floatfmt = self.floatfmt
        except AttributeError:
            if floatfmt is None: floatfmt = 'f'
                
        if colnames==None:
            #colnames = list(self.keys().keys()) # this will prevent byg in Python3 since {}.keys() return dict_keys object instead of a list
            allkeys = list(self.keys().keys())
            colnames = self.order + list(set(allkeys)-set(self.order))
        elif type(colnames) is str:
            colnames = colnames.split()

        def in_notebook():
            # http://stackoverflow.com/questions/15411967/how-can-i-check-if-code-is-executed-in-the-ipython-notebook            
            """
            Returns ``True`` if the module is running in IPython kernel,
            ``False`` if in IPython shell or other Python shell.
            """
            return 'ipykernel' in sys.modules
        
        from tabulate import tabulate as tab # will make my own tabulator in the future
        data = self.getcols(colnames,IDs=IDs,mode=mode,functions=functions)  
        if file:
            with open(file,'w') as f:
                tabstring = tab(zip(*data),colnames,tablefmt=fmt,floatfmt=floatfmt)
                f.write(tabstring)
        else:
            if in_notebook():
                from IPython.core.display import display, HTML
                tabstring = tab(zip(*data),colnames,tablefmt='html',floatfmt=floatfmt)
                display(HTML(tabstring))
            else:
                tabstring = tab(zip(*data),colnames,tablefmt=fmt,floatfmt=floatfmt)
                if raw:
                    return tabstring
                else:
                    print(tabstring)
                #return TabObject(tabstring)
                
    def tabulate_latex(self,caption,label,colnames=None,IDs=None,widths=None,document=None,File=None):
        """
        Tabulate collection in the LaTeX format.
        Parameters:
            widths - list/tuple of relative column widths, should sum up to 1
            document - document class string (e.g. 'article'); if None, then no document enclosement
            File - filename; if None, output is printed to stdout
              
        """
            
        #widths = ''

        TEMPLATE = """\\begin{{table}}[]
\\begin{{adjustbox}}{{width=\\textwidth}}
\\begin{{tabular}}{{{SPEC}}}
\\hline
{HEADER}
\\hline
{BODY}
\\hline
\\end{{tabular}}
\\end{{adjustbox}}
\\caption{{{CAPTION}}}
\\label{{{LABEL}}}
\\end{{table}}
"""
            
        to_str = lambda a: str(a) if a is not None else ''
            
        def to_latex(val):    
            buf = to_str(val).strip()
            # Check if value is a math expression.
            if len(buf)>0 and buf[0]==buf[-1]=='$':
                return buf
            # escape underscores
            buf = buf.replace('_','\_')
            return buf
            
        assert self.order # must be non-empty and not None
            
        # Initialize parameters.
        if IDs is None: IDs = self.ids()
        if colnames is None: colnames = self.order
                    
        # Get table spec depending on widths list.
        if widths is not None:
            SPEC = ''.join(['p{%f\\linewidth}'%w for w in widths])
        else:
            SPEC = ''.join(['l' for k in colnames])
            
        # Get header.
        HEADER = ' & '.join(['\\textbf{%s}'%to_latex(key) for key in colnames]) + '\\\\'

        # Get body.
        LINES = []
        for item in self.subset(IDs).getitems():
            LINE = ' & '.join([to_latex(item.get(key)) for key in colnames]) + '\\\\'
            LINES.append(LINE)
        BODY = '\n'.join(LINES)
            
        # Get the full table LaTex snippet.
        LATEX = TEMPLATE.format(SPEC=SPEC,HEADER=HEADER,
            BODY=BODY,CAPTION=caption,LABEL=label)
        
        if document is not None:
            LATEX = '\\documentclass{%s}\n\\usepackage{adjustbox}\n\\begin{document}\n%s\\end{document}'%\
                (document,LATEX)
        
        # Print or save to file.
        if File is None:
            print(LATEX)
        else:
            with open(File,'w') as f:
                f.write(LATEX)
                
    def head(self,colnames=None,limit=10,fmt='simple',
            functions=None,raw=False,floatfmt=None):
        self.tabulate(colnames=colnames,IDs=self.ids()[:limit],
            fmt=fmt,functions=functions,raw=raw,floatfmt=floatfmt)

    def tail(self,colnames=None,limit=10):
        ids = self.ids()
        self.tabulate(colnames=colnames,IDs=ids[len(ids)-limit:len(ids)])
    
    # old WORKING version
    def update(self,items,IDs=None):
        if type(items) is dict:
            items = [items]
        elif type(items) not in [list,tuple]:
            raise Exception('Items should be either list or tuple')
        if not IDs:
            IDs = self.getfreeids(len(items))
        #if type(IDs) is int:
        #    IDs = [IDs]
        elif type(IDs) is not list:
            raise Exception('Wrong IDs type: %s (expected list or integer)'%type(IDs))
        for ID,item in zip(IDs,items):
            if ID not in self.__dicthash__:
                self.__dicthash__[ID] = {}
            self.__dicthash__[ID].update(item)

    # new unreliable version
    #def update(self,items,merge=False): # this version assumes IDs are in items
    #    # if merge is True, new item is blended into existing item with the same id (if present)
    #    # if merge is False, new item overwrites existing item with the same id
    #    #if type(items) is dict:
    #    if issubclass(items.__class__,dict):
    #        items = [items]
    #    elif type(items) not in [list,tuple]:
    #        raise Exception('Items should be either list or tuple of dict children')
    #    IDs = self.getfreeids(len(items))
    #    for item in items:
    #        if '__id__' not in item:
    #            ID = IDs.pop(0)
    #            item['__id__'] = ID
    #        else:
    #            ID = item['__id__']
    #        if merge:
    #            if ID not in self.__dicthash__:
    #                self.__dicthash__[ID] = {}
    #            self.__dicthash__[ID].update(item)
    #        else:
    #            self.__dicthash__[ID] = item
        
    def delete(self,IDs):
        for ID in IDs:
            del self.__dicthash__[ID]
        
    def group(self,expr): # TODO: add process_exp
        buffer = {}
        if type(expr)==str:
            #expr_ = eval('lambda var: var["' + expr + '"]') # simple
            colnames = expr.split()
            if len(colnames)==1:
                expr_ = lambda v: v[colnames[0]]
            else:
                expr_ = lambda v: tuple([v[k] for k in colnames])
        elif type(expr) in {list,tuple}:
            expr_ = lambda v: tuple([v[k] for k in expr])
        else:
            expr_ = expr
        for ID in self.__dicthash__:
            var = self.__dicthash__[ID] 
            group_value = expr_(var)
            if group_value not in buffer:
                buffer[group_value] = []
            buffer[group_value].append(ID)
        return buffer
        
    def stat(self,keynames,grpi,valname,map=None,reduce=None,plain=False):  # Taken from Jeanny v.4 with some changes
        """
        Calculate function on index values.
        User must provide:
            -> group index (grpi)
            -> mapper and reducer functions.
        MAPPER: item->value (can be scalar or vector)
        REDUCER: item_dict_array->value (can be scalar or vector)
        Flat: True - return plain stat index, False - return Collection
        """
        if map is None: map = lambda v: v
        if reduce is None: reduce = lambda ee: ee
        group_buffer = grpi
        stat_index = {}
        for index_id in group_buffer:
            ids = group_buffer[index_id]
            items = self.getitems(ids)
            map_values = [map(item) for item in items]
            reduce_value = reduce(map_values)
            stat_index[index_id] = reduce_value
        if plain: # return only stat index
            return stat_index 
        else: # return index-based collection
            if type(keynames) is str:
                keynames = keynames.split()
            elif type(keynames) not in {tuple,list}:
                raise Exception('keynames must be str, tuple, or list')
            nkeys = len(keynames)
            col = Collection()
            #col.__dicthash__ = {
            #    key:{keyname:key,valname:stat_index[key]} \
            #    for key in stat_index}
            for keyvals in stat_index:
                if nkeys==1: 
                    keyvals_ = (keyvals,)
                else:
                    keyvals_ = keyvals
                item = {}
                for keyname,keyval in zip(keynames,keyvals_):
                    item[keyname] = keyval
                item[valname] = stat_index[keyvals]
                col.__dicthash__[keyvals] = item
            col.order = keynames + [valname,]
            return col
            
    def stat_(self,keynames,grpi,map=None,reduce=None):
        """
        Calculate function on index values.
        User must provide:
            -> group index (grpi)
            -> mapper and reducer functions.
        MAPPER: item->value (can be scalar or vector)
        REDUCER: item_dict_array->value (can be scalar or vector)
        MAPPER and REDUCER are dicts with the same set of keys,
           each of which is the equivalent of the valname.
           If some keys are absent in mapper, the default 
           mapping is used for this statkey.
        Flat: True - return plain stat index, False - return Collection
        """
        if map is None: map = {}
        if reduce is None: reduce = {}
        
        # Do assertions.
        assert type(map) is dict and map # map should be a non-empty dict
        assert type(reduce) is dict and reduce # reduce should be a non-empty dict
        
        # Get the sequence of keys for statistics
        statkeys = list(reduce.keys())
                
        # Make first table.
        stat = self.stat(
            keynames,grpi,
            statkeys[0],
            map = map.get(statkeys[0]),
            reduce = reduce[statkeys[0]],
        )
        
        # Loop over other stats.
        for statkey in statkeys[1:]:
            stat.join(
                keynames,
                self.stat(
                    keynames,grpi,
                    statkey,
                    map = map.get(statkey),
                    reduce = reduce[statkey],
                ),
                [statkey]
            )
        
        return stat
    
    def sort(self,colnames,IDs=-1,strict=True,mode='greedy',functions=None): # switched default to "greedy" instead of "strict"
        """
        Return permutation of the input IDs.
        """
        if IDs==-1:
            IDs = self.ids()
        vals = self.getcols(colnames=colnames,IDs=IDs,strict=strict,mode=mode,functions=functions)
        vals = [list(e)+[id] for e,id in zip(zip(*vals),IDs)]
        IDs_res = [e[-1] for e in sorted(vals)]
        return IDs_res

    def join(self,key,col,colnames=None,prefix='',strict=True):
        """
        Join column set of external collection (colnames)
        to self using the key, assuming the following conditions:
        1) Self must contain the column named as key.
        2) Col must have the corresponding values of key in its index.
        Key can be either lambda function on item, or a field name
        If strict==False, then ignore if colname is not in col item.
        """
        if type(key)==str:
            key_ = lambda v: v[key]
        elif type(key) in {tuple,list}:
            key_ = lambda v: tuple([v[k] for k in key])
        else:
            key_ = key # expecting lambda func, returning col's index values
        if colnames is None:
            colnames_empty = True; order_set = set()
        else:
            colnames_empty = False        
        for item in self.getitems():
            k_ = key_(item)
            if k_ not in col.__dicthash__:
                continue # skip to next iteration if key is not found in external col
            external_item = col.__dicthash__[k_]            
            if colnames_empty:
                colnames = list(external_item.keys())
            for colname in colnames:
                colname_ = prefix+colname
                if colname_ in item:
                    raise Exception('Collection item already has "%s" field'%colname)
                if not strict and colname not in external_item:
                    continue
                item[colname_] = external_item[colname]
                if colnames_empty: order_set.add(colname_)
        if colnames_empty: 
            self.order += order_set
        else:
            self.order += [prefix+colname for colname in colnames]
        
    def deduplicate(self,colnames=None):
        """
        Deduplicate collection using optional list of colnames.
        If colnames are omitted, all values are used in comparisons.
        
        ATTENTION: in case when colnames is a subset of all columns,
                   there can be a loss of data because other columns
                   can be different. Generally, in this case the algorithm
                   takes the first item in series, and discards other ones.
        """
        
        def subdict(dct): 
            if colnames is None: return dct
            return {key:dct.get(key) for key in colnames}
        
        #def equal(dct1,dct2):
        #    if colnames is None:
        #        return dct1==dct2
        #    else:
        #        return subdict(dct1)==subdict(dct2)
                        
        def to_tuple(dct):
            dct_ = subdict(dct)
            return tuple(zip(dct_.keys(),dct_.values()))
            
        dicthash = self.__dicthash__
        ids_sorted = sorted(dicthash,key=lambda id_: to_tuple(dicthash[id_]))
        
        ids_dedup = []
        for id_ in ids_sorted:
            if len(ids_dedup)==0:
                ids_dedup.append(id_)
            else:
                current_item = dicthash[id_]
                previous_item = dicthash[ids_dedup[-1]]
                if to_tuple(current_item)!=to_tuple(previous_item):
                    ids_dedup.append(id_)
            
        return self.subset(ids_dedup)
        
    # =======================================================
    # =================== UNROLL/UNWIND =====================
    # =======================================================
    def unroll(self,keys,IDs=-1):
        """
        Perform the unrolling of the iterable field(s).
        For example, for {'a':[1,2,3],'b':[10,20]} this will give a new collection:
        {'a':1,'b':10}
        {'a':1,'b':20}
        {'a':2,'b':10}
        {'a':2,'b':20}
        {'a':3,'b':10}
        {'a':3,'b':20}        
        The items not containing the keys from the list will be intact. 
        This method does nearly the same as the "unwind" in MongoDB         
        """
        if type(keys) not in [list,tuple]:
            keys = [keys]
        if IDs==-1:
            IDs = self.ids()
        col = Collection(); col.order = self.order
        def to_list(val):
            if type(val) in [str,unicode]: # string is a special case
                return [val]
            elif type(val) in [list,tuple]: # list and tuple don't need further conversion
                return val
            else: # other cases; TODO: should add iterables separately in the future
                return [val]
        for item in self.getitems(IDs):
            # get the keys from the list which are present in the current item
            active_keys = []
            for key in keys:
                if key in item: active_keys.append(key)
            if not active_keys: # no keys at all
                col.update(item)
            else: # some keys have been found
                for vals in it.product(*[to_list(item[key]) for key in active_keys]):
                    new_item = item.copy()
                    #new_item.update({key:val for key,val in zip(active_keys,vals)}) # this doesn't work in earlier Python versions
                    for key,val in zip(active_keys,vals):
                        new_item[key] = val
                    col.update(new_item)
        return col

    # =======================================================
    # ======================= CSV ===========================
    # =======================================================
    
    def import_csv(self,filename,delimiter=';',quotechar='"',header=None,duck=True): # old_version
        """
        Reads csv-formatted files in more or less robust way.
        Includes avoiding many parsing errors due to "illegal"
        usage of delimiters and quotes.
        """
        # TODO: use csv.Sniffer to deduce the format automatically
        items = []
        with open(filename,'r') as f:
            reader = csv.reader(f,delimiter=delimiter,quotechar=quotechar)
            if not header:                
                colnames = next(reader) # take the first line as header (even if it's really absent)
            else:
                colnames = header            
            nitems = 0
            for vals in reader:
                nitems += 1
                item = {}
                for colname,val in zip(colnames,vals):
                    if val in [None,'']: continue
                    if duck: # duck typing
                        try:
                            val = int(val)
                        except ValueError as e:
                            try:
                                val = float(val)
                            except ValueError as e:
                                if val.strip().lower() in ['f','false','.false.']:
                                    val = False
                                elif val.strip().lower() in ['t','true','.true.']:
                                    val = True
                                else:
                                    pass
                    if type(val) in [str,unicode]: # f..king encoding problems 
                        try:
                            unicode(val)
                        except UnicodeDecodeError:
                            raise Exception('encoding/decoding problems with %s'%val)
                    #if val: item[colname] = val
                    item[colname] = val
                items.append(item)
        self.clear()
        self.update(items)
        self.order = colnames
        return {'nitems':nitems}
        
    def import_csv_(self,filename): # new_version
        """
        Reads csv-formatted files in more or less robust way.
        Includes avoiding many parsing errors due to "illegal"
        usage of delimiters and quotes.
        """
        with open(filename, newline='') as csvfile:
            dialect = csv.Sniffer().sniff(csvfile.read(10000))
            csvfile.seek(0)
            reader = csv.reader(csvfile, dialect)
            header = next(reader)
            items = []
            for vals in reader:
                item = {key:val for key,val in zip(header,vals)}
                items.append(item)
            self.clear()
            self.update(items)
            self.order = header
        
    def export_csv(self,filename,delimiter=';',quotechar='"',order=[]):
        """
        Writes csv-formatted files in more or less robust way.
        Includes avoiding many parsing errors due to "illegal"
        usage of delimiters and quotes.
        Order contains key names which will go first.
        Order saves from reordering columns in Excel each time 
        the CSV file is generated.
        """
        if not order: order = self.order
        keys = self.keys(); 
        header = [key for key in order] + \
                 [key for key in keys if key not in order] # ordered keys must go first
        with open(filename,'w') as f:
            writer = csv.writer(f,delimiter=delimiter,
                  quotechar=quotechar,lineterminator='\n',
                  quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            for ID in self.__dicthash__:
                item = self.__dicthash__[ID]
                vals = []
                for colname in header:
                    if colname in item and item[colname] is not None:
                        vals.append(unicode(item[colname]))
                    else:
                        vals.append('')
                writer.writerow(vals)
                
    # =======================================================
    # ======================= xlsx ==========================
    # =======================================================
    
    def import_xlsx(self,filename):
        """
        Read in the table stored in the Excel file.
        The upper row must contain the column names.
        """
        from openpyxl import Workbook,load_workbook

        # read workbook
        #wb = load_workbook(filename,use_iterators=True) # Python 3: TypeError: load_workbook() got an unexpected keyword argument 'use_iterators'
        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        rowlist = list(sheet) # can be inefficient when the file is large; better to use iterators in this case

        # get row and column count
        #http://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
        #row_count = sheet.max_row
        #column_count = sheet.max_column
    
        # get header, i.e. names of the columns
        header = [cell.value for cell in rowlist[0]]
    
        # fill collection with data
        items = []
        nitems = 0
        for row in rowlist[1:]:
            values = [cell.value for cell in row]
            nitems += 1
            item = {}
            for colname,val in zip(header,values):
                #if val: item[colname] = val # !!!!! BUG: if val==0, it will not be recorded!!!
                if val is not None: item[colname] = val 
            #items.append(item)
            # !!!!!! BUG IN EXCEL/openpyxl: if rows are deleted in Excel, they are still there in the sheet, which will result in many empty items
            if item: items.append(item) 

        self.clear()
        self.update(items)
        return {'nitems':nitems}
        
    # =======================================================
    # ===================== Folder ==========================
    # =======================================================
            
    def import_folder(self,dirname,regex='\.json$'):
        FILENAME_ID = SETTINGS['FILENAME_ID']
        filenames = scanfiles(dirname,regex)
        items = []
        for filename in filenames:
            with open(os.path.join(dirname,filename)) as f:
                try:
                    item = json.load(f)
                except:
                    print('ERROR: %s'%filename)
                    raise
                #if FILENAME_ID in item:
                #    raise Exception('%s has a key %s'%(filenames,FILENAME_ID))
                item[FILENAME_ID] = filename
                items.append(item)
        self.clear()
        self.update(items)
    
    def export_folder(self,dirname,ext='json',default=json_serial):
        """
        Updated version with integrity checking
        to prevent overwriting items in the folder
        in the case when there are similar file names.
        """
        FILENAME_ID = SETTINGS['FILENAME_ID']
        ITEM_ID = SETTINGS['ITEM_ID']
        if not os.path.isdir(dirname):
            #os.mkdir(dirname) # this doesn't work if there are sub-folders
            os.makedirs(dirname)
        # prepare file name index
        FILENAMES = {}
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            if FILENAME_ID in item:
                filename = item[FILENAME_ID]
                if ext:
                    filename,_ = os.path.splitext(filename)
                    filename += ext if ext[0]=='.' else '.'+ext
            elif ITEM_ID in item:
                filename = item[ITEM_ID] 
                filename += ext if ext[0]=='.' else '.'+ext
            else:
                #filename = str(uuid.UUID(bytes=os.urandom(16),version=4)) 
                filename = uuid() 
                filename += ext if ext[0]=='.' else '.'+ext
            FILENAMES[ID] = filename
        # check this index for integrity
        #if len(set(FILENAMES.keys()))!=len(FILENAMES.keys()):# BUG!!!!!!
        if len(set(FILENAMES.values()))!=len(FILENAMES.values()):
            raise Exception('%s index is not unique'%FILENAME_ID)
        # if everything is OK save the collection
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID].copy()
            if FILENAME_ID in item:
                del item[FILENAME_ID]
            filename = FILENAMES[ID]
            with open(os.path.join(dirname,filename),'w') as f:
                #json.dump(item,f,indent=2) # default "dumper"
                json.dump(item,f,indent=2,default=default) # custom "dumper"
                     
    def update_folder(self,dirname,regex=''):
        """
        Update the Jeanny folder collection with self.
        !!! BOTH COLLECTIONS MUST HAVE THE SAME IDS !!!
        !!! ONE MUST USE THE FILENAME_ID FOR THESE PARAMETERS !!!
        """
        FILENAME_ID = SETTINGS['FILENAME_ID']
        dest_col = Collection()
        if os.path.isdir(dirname):
            dest_col.import_folder(dirname,regex)
            dest_col.index('var["%s"]'%FILENAME_ID)
        for item in self.getitems():
            FILENAME = item[FILENAME_ID]
            if FILENAME in dest_col.__dicthash__.keys():
                dest_item = dest_col.getitem(FILENAME)
                dest_item.update(item)
            else:
                dest_col.update(item)
        dest_col.export_folder(dirname)

    # =======================================================
    # ===================== JSON List =======================
    # =======================================================
        
    def import_json_list(self,filename,id=None):
        with open(filename,'r') as f:
            buffer = json.load(f)
        items = []
        for item in buffer:
            items.append(item)
        self.clear()
        self.update(items)
        
    def export_json_list(self,filename,default=json_serial):
        buffer = self.getitems(self.ids())
        with open(filename,'w') as f:
            #json.dump(buffer,f,indent=2) # default "dumper"
            json.dump(buffer,f,indent=2,default=default) # custom "dumper"

    # =======================================================
    # ===================== JSON Dicthash ===================
    # =======================================================
                    
    def export_json_dicthash(self,filename,default=json_serial):
        with open(filename,'w') as f:
            #json.dump(self.__dicthash__,f,indent=2) # default "dumper"
            json.dump(self.__dicthash__,f,indent=2,default=default) # custom "dumper"
            
    # =======================================================
    # ============= Fixcol/Parse from string ================
    # =======================================================
            
    def import_fixcol(self,filename,ignore=True,substitute=None):
        """
        Create collection from the specially formatted column-fixed file.
        THe file must be supplied in the following format (type can be omitted):
        
        //HEADER
        0 Column0 Type0
        1 Column1 Type1
        ...
        N ColumnN TypeN
        
        //DATA
        0___1___2____.....N______
        .... data goes here ....
        
        Comments are marked with hashtag (#) and ignored.
        
        If ignore set to False, exception is thrown at any conversion problems.
        """               
        TYPES = {'float':float,'int':int,'str':str}
        
        f = open(filename)
    
        # Search for //HEADER section.    
        for line in f:
            if '//HEADER' in line: break
    
        # Scan //HEADER section.     
        HEAD = {}        
        for line in f:
            line = line.strip()
            if not line: continue
            if line[0]=='#': continue
            if '//DATA' in line: break
            vals = [_ for _ in line.split() if _]
            token = vals[0]
            if token in HEAD:
                raise Exception('ERROR: duplicate key was found: %s'%vals[0])
            vtype = TYPES[vals[2]] if len(vals)>2 else str           
            HEAD[token] = {}
            HEAD[token]['token'] = token
            HEAD[token]['name'] = vals[1]
            HEAD[token]['type'] = vtype # vtype
                        
        # Get tokenized mark-up.
        for line in f:
            widths = line.rstrip(); break # readline doesn't work because of the "Mixing iteration and read methods"
        matches = re.finditer('([^_]+_*)',widths)
        tokens = []; names = []
        for match in matches:
            i_start = match.start()
            i_end = match.end()
            token = re.sub('_','',widths[i_start:i_end])
            if token not in HEAD: continue                
            tokens.append(token)
            names.append(HEAD[token]['name'])
            HEAD[token]['i_start'] = i_start
            HEAD[token]['i_end'] = i_end
        #markup = re.findall('([^_]+_*)',widths) # doesn't give indexes
        
        # Scan //DATA section.     
        items = []
        for line in f:
            if line.strip()=='': continue
            if line.lstrip()[0]=='#': continue
            line = line.rstrip()
            #item = {HEAD[token]['name']:HEAD[token]['type'](line[HEAD[token]['i_start']:HEAD[token]['i_end']]) for token in HEAD} # doesn't work in earlier versions of Python
            item = {}
            for token in HEAD:
                try:
                    buf = line[HEAD[token]['i_start']:HEAD[token]['i_end']]
                    val = HEAD[token]['type'](buf)
                except ValueError as e:
                    if not ignore:
                        raise Exception(e)
                    else:
                        val = substitute
                item[HEAD[token]['name']] = val
            items.append(item) 
        self.clear()
        self.setorder(names)
        self.update(items)    
    
    def export_fixcol(self,filename):
        """ 
            ATTENTION:
            This was done to be used ONLY if a column-fixed format is mandatory!!!
            This function is buggy and will cause non-reversible changes in the 
            the string values of teh items.
            Use other export formats to achieve stability. 
        """
        
        # conversion to string accounting for None values
        def to_str(val):
            if val is not None:
                return str(val)
            else:
                return ''
        # get order
        order = self.order.copy()
        order += list( set(self.keys())-set(order) )
        # deduce types from the collection
        checked = set()
        types = {}
        ids = self.ids()
        for id_ in ids:
            item = self.getitem(id_)
            if not set(order)-checked: break
            for colname in order:
                if colname in item and item[colname] is not None:
                    types[colname] = type(item[colname])
                    checked.add(colname)
        # if items are not present, assign str type for them
        for colname in set(order)-checked:
            types[colname] = str
        with open(filename,'w') as f:
            f.write('//HEADER\n')
            # make a header
            tokens = string.digits+string.ascii_uppercase+string.ascii_lowercase
            tokens = tokens[:len(order)]
            for token,colname in zip(tokens,order):
                f.write('%s %s %s\n'%(token,colname,types[colname].__name__))
            f.write('\n//DATA\n')
            # get columns and find widths
            COLS = self.getcols(order)
            COLS_STR = [[to_str(e) for e in col] for col in COLS]
            get_width = lambda col: max([len(s) for s in col])
            widths = [get_width(col) for col in COLS_STR]
            # write tokenized header
            dw = 3 # gap between columns
            for token,width in zip(tokens,widths):
                f.write(token+'_'*(width-1+dw))
            f.write('\n')
            # write the content, do conversion checks
            ncols = len(COLS_STR)
            for i in range(len(ids)):
                for j in range(ncols):
                    type_ = types[order[j]]
                    strval = COLS_STR[j][i]
                    trueval = COLS[j][i]
                    if trueval is not None: 
                        # this should signalize if there are conversion problems
                        if trueval!=type_(strval):
                            raise Exception('conversion error: ',trueval,type_(strval)) 
                    width = widths[j]
                    f.write('%%%ds'%(width+dw)%strval)
                f.write('\n')
    
    # =======================================================
    # ============= XSCDB/HAPI2.0 STUFF =====================
    # =======================================================
    
    def xscdb_lookup_molecule(self,colname,IDs=-1):
        import xscdb
        if not xscdb.VARSPACE['session']:
            xscdb.start()
        lookup = lambda al: xscdb.query(xscdb.Molecule).\
                            join(xscdb.MoleculeAlias).\
                            filter(xscdb.MoleculeAlias.alias.like(al)).first()
        if IDs==-1: 
            IDs = self.ids()
        for ID in IDs:
            v = self.__dicthash__[ID]
            altypes = ['name','csid','cas','acronym']
            res = {}
            for altype in altypes:                
                if altype in v:
                    dbitem = lookup(v[altype])
                    if dbitem is not None: res[altype] = dbitem
            vals = set(res.values())
            if len(vals)>1:
                #print('WARNING: more than one entrie found for %s: %s'%\
                #  ({altype:v[altype] for altype in altypes if altype in v},res)) # doesn't work in earlier versions of Python
                aa = {}
                for altype in altypes:
                    if altype in v: aa[altype] = v[altype]
                print('WARNING: more than one entrie found for %s: %s'%(aa,res))
                v[colname] = tuple(vals)
            else:
                v[colname] = tuple(vals)[0]
                
    # =======================================================
    # ============= Checksums and integrity =================
    # =======================================================
    
    def md5(self,v):# Jeanny4: this must belong to item, not to collection
        return hashlib.md5('%s'%[v[key] for key in sorted(v.keys())]).hexdigest()

    # =======================================================
    # ============= Typing... ===============================
    # =======================================================
    
    def get_types(self,nitems=None):
        types = {}
        dicthash = self.__dicthash__
        if nitems is None:
            nitems = len(dicthash)
        for i,c in zip(range(nitems),dicthash):
            item = dicthash[c]
            tt = {key:type(item[key]) for key in item}
            for key in tt:
                ty = tt[key]
                if key not in types:
                    types[key] = ty
                elif types[key] is None:
                    types[key] = ty
        return types

    # =======================================================
    # ============= Equality... =============================
    # =======================================================
        
    def __eq__(self,other):
        return self.types==other.types and self.__dicthash__==other.__dicthash__
        
    def __ne__(self,other):
        return not self==other

    # =======================================================
    # ============= Plotting... =============================
    # =======================================================

    def plotlayers_lineseries(
                    self,
                    xkey,ykey,
                    name=None,
                    markerstyle=None,
                    markersize=None,
                    linestyle=None,
                    color=None,
                    functions={},
                    options={}
                    ):
        """
        xkey - list of keys of the Y axes (can be a single string key)
        ykey - list of keys of the Y axes (can be a single string key)
        
        name - list of legend names, or a single name
        
        {markerstyle,markersize,linestyle,color,...} - specs
        Each spec is either None, or string, or list of strings
        
        functions - dictionary of auxiliary lambda functions.
        
        options - dictionary Layer of options.
        
        This method returns a list of layers of type LineSeries and datatype DataPoints2D.
        """
        
        if type(ykey) is str:
            ykey = [ykey]

        n_series = len(ykey)
        if type(xkey) is str:
            xkey = [xkey]*n_series
            
        assert len(xkey)==len(ykey)
        
        for yk in ykey:
            assert type(yk) is str # ykey must be str! otherwise use "functions" argument
        
        columns = self.getcols(xkey+ykey,functions=functions,mode='greedy')
        
        x_columns = columns[:n_series]
        y_columns = columns[n_series:]
        
        dtype = DataPoints2D
        ltype = LineSeries
        lopts = LineSeriesOptions(**options)
        lopts['line_style'] = ''
                
        def process(spec):        
            if spec is None or type(spec) in {str,int,float}:
                spec = [spec]*len(ykey)
            else:
                spec = cycle(spec)
            return spec
        
        markerstyle = process(markerstyle)
        markersize = process(markersize)
        linestyle = process(linestyle)
        color = process(color)
        name = process(name)
        
        layers = []

        for xcol,ycol,yk,nm,mrkst,mrksz,lnst,clr in zip(
                x_columns,
                y_columns,
                ykey,
                name,
                markerstyle,
                markersize,
                linestyle,
                color,
        ):
            
            rows = [(x,y) for x,y in zip(xcol,ycol) if x is not None and y is not None]
            if not rows: continue
            
            x,y = list(zip(*rows))
            data = dtype(x,y)
            
            lopts_ = lopts.copy()
            if mrkst is not None: lopts_['marker_style'] = mrkst
            if mrksz is not None: lopts_['marker_size'] = mrksz
            if lnst is not None: lopts_['line_style'] = lnst
            if clr is not None: lopts_['line_color'] = clr
            if clr is not None: lopts_['marker_color'] = clr
                
            #lopts_['line_style'] = ''
            #lopts_['marker_style'] = 'o'
            #lopts_['marker_size'] = 20
                
            if not nm: nm = yk
            
            layer = ltype(data,nm,lopts_)
            layers.append(layer)
        
        return layers
    
    def plot(
                self,
                xkey,ykey,
                xlabel=None,
                ylabel=None,
                name=None,
                title=None,
                markerstyle=None,
                markersize=None,
                linestyle=None,
                color=None,
                logscale_x=False,
                logscale_y=False,
                xlim=None,
                ylim=None,
                size=False,
                functions={},
                axes_options={},
                layer_options={},
                ):
        """
        Simple plot of lineseries.
        """
        
        # Create Axes options.
        aopts = AxesOptions(**axes_options)
        if logscale_x: aopts['x_axis_logscale_on'] = True
        if logscale_y: aopts['y_axis_logscale_on'] = True
        if title: aopts['title'] = title
        if xlabel: aopts['x_axis_label'] = xlabel
        if ylabel: aopts['y_axis_label'] = ylabel
        if xlim: aopts['x_axis_limits'] = xlim
        if ylim: aopts['y_axis_limits'] = ylim
            
        # Create plot options.
        popts = {}
        if size: popts['size'] = size
                
        # Create layers.
        layers = self.plotlayers_lineseries(
            xkey,ykey,
            name=name,
            markerstyle=markerstyle,
            markersize=markersize,
            linestyle=linestyle,
            color=color,
            functions=functions,
            options=layer_options,
        )
        
        ax = Axes(layers,options=aopts)
        ax.plot(size=size)
        
    def plotlayers_errorbars(
                    self,
                    xkey,ykey,yerrkey,
                    xerrkey=None,
                    name=None,
                    markerstyle=None,
                    markersize=None,
                    linestyle=None,
                    color=None,
                    functions={},
                    options={}
                    ):
        """
        xkeys - list of keys of the Y axes (can be a single string key)
        ykeys - list of keys of the Y axes (can be a single string key)
        yerrkeys - list of keys of the Y errorbar axes (can be a single string key)
        xerrkeys - list of keys of the X errorbar axes (can be a single string key or None)
        
        {markerstyle,markersize,linestyle,color,...} - specs
        Each spec is either None, or string, or list of strings
        
        functions - dictionary of auxiliary lambda functions.
        
        options - dictionary Layer of options.
        
        This method returns a list of layers of type Errorbars and datatype DataErrorbars.
        """
        
        if type(ykey) is str:
            ykey = [ykey]

        n_series = len(ykey)
        if type(xkey) is str:
            xkey = [xkey]*n_series
        
        if type(yerrkey) is str:
            yerrkey = [yerrkey]*n_series
        
        xerrkey = [xerrkey]*n_series if type(xerrkey) is str else None
        
        if xerrkey is None:
            assert len(xkey)==len(ykey)==len(yerrkey)
        else:
            assert len(xkey)==len(ykey)==len(yerrkey)==len(xerrkey)
        
        for i in range(n_series):
            assert type(xkey[i]) is str
            assert type(ykey[i]) is str
            assert type(yerrkey[i]) is str
            if xerrkey is not None: assert type(xerrkey[i]) is str
        
        columns = self.getcols(xkey+ykey+yerrkey,functions=functions,mode='greedy')
        
        x_columns = columns[:n_series]
        y_columns = columns[n_series:2*n_series]
        yerr_columns = columns[2*n_series:]
        
        assert len(x_columns)==len(y_columns)==len(yerr_columns)==n_series # debug
        
        if xerrkey is not None:
            xerr_columns = self.getcols(xerrkey,functions=functions,mode='greedy')
        else:
            xerr_columns = [None]*n_series
        
        dtype = DataErrorbars
        ltype = Errorbars
        lopts = ErrorbarsOptions(**options)
        lopts['line_style'] = ''
                
        def process(spec):        
            if spec is None or type(spec) in {str,int,float}:
                spec = [spec]*len(ykey)
            else:
                spec = cycle(spec)
            return spec
        
        markerstyle = process(markerstyle)
        markersize = process(markersize)
        linestyle = process(linestyle)
        color = process(color)
        name = process(name)
        
        layers = []

        for xcol,ycol,yerr_col,xerr_col,yk,nm,mrkst,mrksz,lnst,clr in zip(
                x_columns,
                y_columns,
                yerr_columns,
                xerr_columns,
                ykey,
                name,
                markerstyle,
                markersize,
                linestyle,
                color,
        ):
            
            if xerr_col is None: # X errorbars are omitted
                rows = [(x,y,yerr) for x,y,yerr in zip(xcol,ycol,yerr_col) \
                        if x is not None and y is not None and yerr is not None]
                if not rows: continue            
                x,y,yerr = list(zip(*rows))
                data = dtype(x,y,yerr)
            else: # X errorbars are included
                rows = [(x,y,yerr,xerr) for x,y,yerr,xerr in zip(xcol,ycol,yerr_col,xerr_col) \
                        if x is not None and y is not None and yerr is not None and xerr is not None]
                if not rows: continue            
                x,y,yerr,xerr = list(zip(*rows))
                data = dtype(x,y,yerr,xerr)
            
            lopts_ = lopts.copy()
            if mrkst is not None: lopts_['marker_style'] = mrkst
            if mrksz is not None: lopts_['marker_size'] = mrksz
            if lnst is not None: lopts_['line_style'] = lnst
            if clr is not None: lopts_['line_color'] = clr
            if clr is not None: lopts_['marker_color'] = clr
            if clr is not None: lopts_['errorbar_color'] = clr    
                                
            if not nm: nm = yk
            
            layer = ltype(data,nm,lopts_)
            layers.append(layer)
        
        return layers

    def plot_errorbars(
                self,
                xkey,ykey,yerrkey,
                xerrkey=None,                
                xlabel=None,
                ylabel=None,
                name=None,
                title=None,
                markerstyle=None,
                markersize=None,
                linestyle=None,
                color=None,
                logscale_x=False,
                logscale_y=False,
                xlim=None,
                ylim=None,
                size=False,
                functions={},
                axes_options={},
                layer_options={},
                ):
        """
        Plot of lineseries with errorbars.
        """
        
        # Create Axes options.
        aopts = AxesOptions(**axes_options)
        if logscale_x: aopts['x_axis_logscale_on'] = True
        if logscale_y: aopts['y_axis_logscale_on'] = True
        if title: aopts['title'] = title
        if xlabel: aopts['x_axis_label'] = xlabel
        if ylabel: aopts['y_axis_label'] = ylabel
        if xlim: aopts['x_axis_limits'] = xlim
        if ylim: aopts['y_axis_limits'] = ylim
            
        # Create plot options.
        popts = {}
        if size: popts['size'] = size
                
        # Create layers.
        layers = self.plotlayers_errorbars(
            xkey,ykey,
            yerrkey,xerrkey,
            name=name,
            markerstyle=markerstyle,
            markersize=markersize,
            linestyle=linestyle,
            color=color,
            functions=functions,
            options=layer_options,
        )
        
        ax = Axes(layers,options=aopts)
        ax.plot(size=size)
        
    def plotlayers_text(
                    self,
                    xkey,ykey,textkey,
                    name=None,
                    font_color=None,
                    font_size = None,
                    functions={},
                    options={}
                    ):
        """
        xkey - key of the X axis
        ykey - key of the Y axes
        textkey - key of the text axes
        
        {color,...} - specs
        Each spec is either None, or string, or list of strings
        
        functions - dictionary of auxiliary lambda functions.
        
        options - dictionary of options for Text layer.
        
        This method layers of type Text and datatype DataText.
        """
        
        assert type(ykey) is str # ykey must be str! otherwise use "functions" argument
        assert type(textkey) is str # textkey must be str! otherwise use "functions" argument
        
        columns = self.getcols([xkey,ykey,textkey],functions=functions,mode='greedy')
        
        dtype = DataText
        ltype = Text
        lopts = TextOptions(**options)
        
        layers = []
        xcol = columns[0]
        ycol = columns[1]
        textcol = columns[2]
                
        rows = [(x,y,txt) for x,y,txt in zip(xcol,ycol,textcol) \
                if x is not None and y is not None and txt is not None]
        
        x,y,txt = list(zip(*rows))
        data = dtype(x,y,txt)
        
        if font_color is not None: lopts['font_color'] = font_color
        if font_size is not None: lopts['font_size'] = font_size
            
        layer = ltype(data,name,lopts)
        
        return [layer]
    
    def plotlayers_fillbetween(
                    self,
                    xminkey,xmaxkey,
                    yminkey,ymaxkey,
                    name,
                    color='grey',
                    alpha=None,
                    functions={},
                    options={}
                    ):
        """
        xminkey,xmaxkey - keys for the ranges on X
        yminkey,ymaxkey - keys for the ranges on Y
        
        {color,alpha,...} - specs
        Each spec is either None, or string, or list of strings
        
        functions - dictionary of auxiliary lambda functions.
        
        options - dictionary of options for Text layer.
        
        This method layers of type FillBetween and datatype DataFillBetween.
        """
        
        assert type(xminkey) is str
        assert type(xmaxkey) is str
        assert type(yminkey) is str
        assert type(ymaxkey) is str
                
        columns = self.getcols([xminkey,xmaxkey,yminkey,ymaxkey],functions=functions,mode='greedy')
        
        dtype = DataFillBetween
        ltype = FillBetween
        lopts = FillBetweenOptions(**options)
        
        xmin_col = columns[0]
        xmax_col = columns[1]
        ymin_col = columns[2]
        ymax_col = columns[3]
                
        rows = [(xmin,xmax,ymin,ymax) for xmin,xmax,ymin,ymax in zip(xmin_col,xmax_col,ymin_col,ymax_col) \
                if xmin is not None and xmax is not None and ymin is not None and ymax is not None]
        
        xmin,xmax,ymin,ymax = list(zip(*rows))
        data = dtype(xmin,xmax,ymin,ymax)
        
        if color is not None: lopts['color'] = color
        if alpha is not None: lopts['alpha'] = alpha
            
        layer = ltype(data,name,lopts)
        
        return [layer]

    # =======================================================
    # ============= Other unfinished stuff... ===============
    # =======================================================
        
    def import_binary(self,filename):
        pass
        
    def export_binary(self,filename):
        pass

class Tree: # can a collection do that ??????????
    """
    A collection tree used especially for the cluster parallelized computations
    (temporary implementation).
    """
    
    def __init__(self,folders=None,type='folder',regex='\.json'):   
        self.__cols__ = []
        if folders is not None:
            self.read(folders,type=type,regex=regex)
            
    def __repr__(self):
        res = ',\n'.join([str(col) for col in self.__cols__])
        return res
        
    def read(self,folders,type='folder',regex='\.json'):
        for path in folders:
            print('reading collection from %s'%path)
            col = Collection(path=path,type=type,regex=regex)
            self.__cols__.append(col)
                
    def write(self,folders=None,type=None):
        # check pathes first
        for col in self.__cols__:
            if col.__path__ in {'',None}:
                raise Exception('Path should be non-empty for col %s'%col)
        # get folders
        if folders is None:
            folders = []
            for col in self.__cols__:
                folders.append(col.__path__)
        for col,path in zip(self.__cols__,folders):
            print('exporting %s to %s'%(col,path))
            col.export(path=path,type=type)
        
    def assign(self,par,expr):
        for col in self.__cols__:
            col.assign(par,expr)
        
    def assign_(self,expr):
        for col in self.__cols__:
            col.assign_(expr)
            
    def delete(self,**argv):
        for col in self.__cols__:
            col.delete(**argv)
            
    def subset(self,**argv):
        t = Tree()        
        for col in self.__cols__:
            t.__cols__.append(col.subset(**argv))
        return t
            
    def union(self):
        res = Collection()
        for col in self.__cols__:
            res.update(col.getitems())
        return res

class Database:
    """
    A class describing the set of heterogenous collections  
    interconnected with each other.
    """
    
    REGISTRY_NAME = '__registry__.csv'
    RELATIONS_NAME = '__relations__.csv'
    
    def __init__(self,root):
        self.__root__ = root
        # import registry
        self.__registry__ = self.__import_if_exists__(self.REGISTRY_NAME)
        self.__registry__.index('name')
        # import relations
        self.__relations__ = self.__import_if_exists__(self.RELATIONS_NAME)
        # import collections
        for item in self.__registry__.getitems():
            name = item['name']
            path = os.path.join(self.REGISTRY_NAME,name)
            col = import_csv(path)
            
    def __import_if_exists__(self,path):
        root = self.__root__
        path = os.path.join(root,path)
        if os.path.isdir(root):
            if not os.path.isfile(path):
                raise Exception('Bad database format: cannot find %s'%path)
            col = import_csv(path)
        else:
            col = Collection()
        return col
        
    def create_collection(self,collection_name):
        assert collection_name!=self.REGISTRY_NAME
        if collection_name in self.__registry__.__dicthash__:
            raise Exception('Collection "%s" already exists'%\
                collection_name)
        col = Collection()
        self.__registry__.__dicthash__[collection_name] = {
            'name':collection_name,'col':col}
        return col
        
    def get_collection(self,collection_name):
        return self.__registry__.__dicthash__[collection_name]
        
    def __getitem__(self,collection_name):
        return self.get_collection(collection_name)
        
    def add_collection(self,col,collection_name):
        assert collection_name!=self.REGISTRY_NAME
        if collection_name in self.__registry__.__dicthash__:
            raise Exception('Collection "%s" already exists'%\
                collection_name)
        self.__registry__.__dicthash__[collection_name] = {
            'name':collection_name,'col':col}      

    def drop_collection(self,collection_name):
        del self.__registry__.__dicthash__[collection_name]
        
    # TODO !!!
    def join(self,name1,name2,key,key2=None,inner=True):
        # get join index
        col1 = self.get_collection(name1)
        col2 = self.get_collection(name2)
        join_index = create_join_index(col1,col2,key,key2,inner)
        # get column naming scheme
        keys1 = col1.keys()
        keys2 = col2.keys()
        keys_common = set(keys1).intersection(keys2)
        names_dict = lambda name,keys: \
            {k:(k if k not in keys_common else '%s.%s'%(name,k)) for k in keys}
        names_dict_1 = names_dict(name1,keys1)
        names_dict_2 = names_dict(name2,keys2)
        order1 = col1.order
        order2 = col2.order
        ordered_keys = lambda order,keys: list(order1)+list(set(keys1)-order1)
        keys_order_1 = ordered_keys(order1,keys1)
        keys_order_2 = ordered_keys(order2,keys2)
        names_order_1 = [names_dict_1[k] for k in keys_order_1]
        names_order_2 = [names_dict_2[k] for k in keys_order_2]
        # create joined collection using the naming scheme
        col = Collection()
        col.order = names_order_1 + names_order_2
        for id1,id2 in join_index:
            item = {}
            if id1 is not None:
                item1 = col1.getitem(id1)
                for c in item1:
                    c_ = names_dict_1(c)
                    item[c_] = item1[c]
            if id2 is not None:
                item2 = col2.getitem(id2)
                for c in item2:
                    c_ = names_dict_2(c)
                    item[c_] = item2[c]
            col.update(item)
        # return the resulting collection
        return col
        
    def save(self):
        # save registry
        registry_path = os.path.join(self.__root__,self.REGISTRY_NAME)
        if not os.path.exists(self.__root__):
            os.mkdir(self.__root__)
        reg = Collection()
        reg.update([{'name':v['name']} for v in self.__registry__.getitems()])
        reg.export_csv(registry_path)
        # save collections
        for item in self.__registry__.getitems():
            col = item['col']
            name = item['name']
            path = os.path.join(self.REGISTRY_NAME,name)
            col.export_csv(path)
            
    def __repr__(self):
        return 'Database("%s")'%self.__root__
            
class JobManager:
    """
        ncores => number of CPUs
        nnodes=1 => number of nodes
        mempcore=1500 => amount of RAM per core
        name='calc' => default job name
        command=None => command to launch
        walltime='24' => walltime in hours
    """
    def __init__(self,ncores=1,nnodes=1,mempcore=1500,
        name='calc',command=None,walltime='24'):
        self.ncores = ncores
        self.nnodes = nnodes
        self.mempcore = mempcore
        self.name = name
        self.command = command
        self.walltime = walltime
            
class JobManagerSGE(JobManager):
    template = """# /bin/sh 
# ----------------Parameters---------------------- #
#$ -S /bin/sh
#$ -pe mthread {ncores}
#$ -l s_cpu={walltime}:00:00
#$ -l mres={mempcore}M
#$ -cwd
#$ -j y
#$ -N {name}
#$ -o {name}.log
#$ -m bea
#
# ----------------Modules------------------------- #
#module load tools/python2.6-x
module load opt-python
module load intel

source ~/.bashrc
#
# ----------------Your Commands------------------- #
#
echo + `date` job $JOB_NAME started in $QUEUE with jobID=$JOB_ID on $HOSTNAME
echo + NSLOTS = $NSLOTS
#
{command}
#
echo = `date` job $JOB_NAME done
    """
    def __repr__(self):
        return self.template.format(
            ncores=self.ncores,
            nnodes=self.nnodes,
            mempcore=self.mempcore,
            name=self.name,
            command=self.command,
            walltime=self.walltime)
            
class JobManagerSlurm(JobManager):
    def __repr__(self):
        pass

            
# ==============================
# SUPPLEMENTARY HELPER FUNCTIONS
# ==============================

# get names of all files in a given folder
def get_filenames(dirname):
    filenames = [entry for entry in os.listdir(dirname) 
                 if os.path.isfile(os.path.join(dirname,entry))]
    return filenames
    
def get_dirnames(dirname):
    dirnames = [entry for entry in os.listdir(dirname) 
                 if os.path.isdir(os.path.join(dirname,entry))]
    return dirnames

# filter string according to supplied regular expression (PCRE)
def filterstr(lst,regex):
    return [entry for entry in lst if re.search(regex,entry)]

# scan folder for files which obey the given regular expression (PCRE)
def scanfiles(dirname='./',regex=''):
    return filterstr(get_filenames(dirname),regex)
scandir = scanfiles # BACKWARDS COMPATIBILITY!!
    
# scan folder for sub-folders which obey the given regular expression (PCRE)
def scandirs(dirname='./',regex=''):
    return filterstr(get_dirnames(dirname),regex)

def copyfile(srcdir,srcnames,destdir,destnames=None):
    if not destnames: destnames = srcnames
    for srname,destname in zip(srcnames,destnames):
        # unlike copy(), copy2() retains file attributes
        shutil.copy2(os.path.join(srcdir,srname),
                    os.path.join(destdir,destname))
    
# convert old HAPI-formatted table to Collection    
def collect_hapi(LOCAL_TABLE_CACHE,TableName):
    lines = Collection()
    for i in range(LOCAL_TABLE_CACHE[TableName]['header']['number_of_rows']):
        line = {}
        for par in LOCAL_TABLE_CACHE[TableName]['data'].keys():
            line[par] = LOCAL_TABLE_CACHE[TableName]['data'][par][i]
        lines.update(line)
    return lines
    
# ATTENTION!!! BETTER VERSIONS FOR FUNCTIONS FOR WORKING WITH .PAR HITRAN FORMAT 
# ARE GIVEN IN D:\work\Activities\HAPI\EXCEL_INTERFACE\dotpar\dotpar_converter.py
    
# import HITRAN .par file into a collection
"""
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
par_line                                                                                                                                                         |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
281 0.000001901 1.298E-36 1.637E-25.05940.103  672.98580.580.000000      0 0 0 0        0 0 0 0   12  6 0A-      12  6 0A+     935430 5 4 2 2 1 0   200.0  200.0 |
281 0.000002071 2.913E-35 6.874E-25.07050.105   48.61680.670.000000      0 0 0 0        0 0 0 0    3  3 0A-       3  3 0A+     945430 5 4 2 2 1 0    56.0   56.0 |
281 0.000005174 5.096E-36 2.830E-24.05620.101  787.77310.570.000000      0 0 0 0        0 0 0 0   13  6 0A-      13  6 0A+     935430 5 4 2 2 1 0   216.0  216.0 |
281 0.000012976 1.636E-35 3.866E-23.05300.100  911.20010.560.000000      0 0 0 0        0 0 0 0   14  6 0A-      14  6 0A+     935430 5 4 2 2 1 0   232.0  232.0 |
281 0.000014477 9.247E-34 1.411E-22.07300.109   84.21490.660.000000      0 0 0 0        0 0 0 0    4  3 0A-       4  3 0A+     945430 5 4 2 2 1 0    72.0   72.0 |
281 0.000030368 4.412E-35 4.336E-22.04990.099 1043.22630.550.000000      0 0 0 0        0 0 0 0   15  6 0A-      15  6 0A+     935430 5 4 2 2 1 0   248.0  248.0 |
281 0.000057843 9.690E-33 5.998E-21.07280.109  128.68900.650.000000      0 0 0 0        0 0 0 0    5  3 0A-       5  3 0A+     945430 5 4 2 2 1 0    88.0   88.0 |
281 0.000067033 1.019E-34 4.113E-21.04690.097 1183.80910.540.000000      0 0 0 0        0 0 0 0   16  6 0A-      16  6 0A+     935430 5 4 2 2 1 0   264.0  264.0 |
281 0.000140651 2.048E-34 3.375E-20.04390.096 1332.90370.530.000000      0 0 0 0        0 0 0 0   17  6 0A-      17  6 0A+     935430 5 4 2 2 1 0   280.0  280.0 |
281 0.000173303 5.666E-32 1.152E-19.07170.109  182.02350.640.000000      0 0 0 0        0 0 0 0    6  3 0A-       6  3 0A+     945430 5 4 2 2 1 0   104.0  104.0 |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
par_line                                                                                                                                                         |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
281 0.000282342 3.628E-34 2.442E-19.04100.094 1490.46320.520.000000      0 0 0 0        0 0 0 0   18  6 0A-      18  6 0A+     835430 5 4 2 2 1 0   296.0  296.0 |
281 0.000545084 5.719E-34 1.580E-18.03820.092 1656.43890.510.000000      0 0 0 0        0 0 0 0   19  6 0A-      19  6 0A+     835430 5 4 2 2 1 0   312.0  312.0 |
281 0.003514511 3.319E-30 3.669E-16.06150.104  483.54750.600.000000      0 0 0 0        0 0 0 0   10  3 0A-      10  3 0A+     845430 5 4 2 2 1 0   168.0  168.0 |
281 0.006136062 5.752E-30 1.627E-15.05840.103  580.84530.590.000000      0 0 0 0        0 0 0 0   11  3 0A-      11  3 0A+     835430 5 4 2 2 1 0   184.0  184.0 |
281 0.010200974 8.732E-30 6.323E-15.05540.102  686.84820.580.000000      0 0 0 0        0 0 0 0   12  3 0A-      12  3 0A+     835430 5 4 2 2 1 0   200.0  200.0 |
281  17.8054720 9.543E-22 2.333E-04.07820.114    8.90430.690.000000      0 0 0 0        0 0 0 0    2  0 0A+       1  0 0A+     845430 5 4 2 2 1 0    40.0   24.0 |
281  17.8061608 7.176E-22 3.500E-04.07960.113    8.37110.690.000000      0 0 0 0        0 0 0 0    2  1 0E        1  1 0E      845430 5 4 2 2 1 0    20.0   12.0 |
281  26.7003452 2.890E-21 8.428E-04.07420.112   26.70980.680.000000      0 0 0 0        0 0 0 0    3  0 0A+       2  0 0A+     845430 5 4 2 2 1 0    56.0   40.0 |
281  26.7013756 2.575E-21 1.498E-03.07580.113   26.17730.680.000000      0 0 0 0        0 0 0 0    3  1 0E        2  1 0E      845430 5 4 2 2 1 0    28.0   20.0 |
281  26.7044710 1.623E-21 9.371E-04.07550.107   24.57810.680.000000      0 0 0 0        0 0 0 0    3  2 0E        2  2 0E      845430 5 4 2 2 1 0    28.0   20.0 |
+------------------------------------------------------------------------------------------------------------------------------------------------------------------+
 11    0.072059 2.043E-30 5.088E-12.09190.391 1922.82910.760.003700          0 1 0          0 1 0  4  2  2        5  1  5      5545533321287120 7     9.0   11.0
 11    0.900086 5.783E-35 1.965E-08.07990.352 5613.36960.57-.002400          0 3 0          0 3 0  8  2  7        7  3  4      4442434432287122 8    51.0   45.0
 11    0.895092 3.600E-28 8.314E-09.08010.412 2129.59910.75-.000300          0 1 0          0 1 0  5  3  2        4  4  1      5546633321287120 8    33.0   27.0
 11    0.865759 5.104E-35 2.755E-08.08180.352 5435.41210.680.005300          0 3 0          0 3 0  6  3  3        7  2  6      4445534432257120 7    13.0   15.0
 11    0.768939 7.600E-37 1.894E-13.07820.434 4030.06980.690.000800          1 0 0          0 0 1  4  3  2        3  3  1      4342434432257122 7    27.0   21.0
 11    0.766500 1.393E-37 1.714E-08.05840.298 6655.62940.51-.009100          0 3 0          0 3 0  9  6  4       10  5  5      3342434432297122 9    19.0   21.0
""";

def dotpar_item_to_list(item):
    elements = ('molec_id','local_iso_id','nu','sw','a','gamma_air','gamma_self','elower','n_air','delta_air',
    'global_upper_quanta','global_lower_quanta','local_upper_quanta','local_lower_quanta','ierr','iref','gp','gpp')
    lst = [item[e] for e in elements]
    return lst

def load_dotpar(line):
    """
    Get the "raw" .par line on input
    and output the dictionary of parameters.
    """
    item = dict(
        molec_id             = int(   line[  0:  2] ),
        local_iso_id         = int(   line[  2:  3], base=16),        
        nu                   = float( line[  3: 15] ),
        sw                   = float( line[ 15: 25] ),
        a                    = float( line[ 25: 35] ),
        gamma_air            = float( line[ 35: 40] ),
        gamma_self           = float( line[ 40: 45] ),
        elower               = float( line[ 45: 55] ),
        n_air                = float( line[ 55: 59] ),
        delta_air            = float( line[ 59: 67] ),
        global_upper_quanta  = str(   line[ 67: 82] ),
        global_lower_quanta  = str(   line[ 82: 97] ),
        local_upper_quanta   = str(   line[ 97:112] ),
        local_lower_quanta   = str(   line[112:127] ),        
        ierr                 = str(   line[127:133] ),        
        iref                 = str(   line[133:145] ),        
        gp                   = float( line[145:153] ),
        gpp                  = float( line[153:160] ),  
    )   
    if item['local_iso_id']==10: item['local_iso_id']=0
    return item
        
def import_dotpar(filename):
    col = Collection()
    with open(filename) as f:
        for line in f:
            item = load_dotpar(line)
            col.update(item)
    col.order = list(item.keys())
    return col

def import_fixcol(filename,*args,**kwargs):
    col = Collection()
    col.import_fixcol(filename,*args,**kwargs)
    return col

def import_csv(filename,*args,**kwargs):
    col = Collection()
    col.import_csv(filename,*args,**kwargs)
    return col

def export_to_hapi_cache(col,table_name,LOCAL_TABLE_CACHE,HITRAN_DEFAULT_HEADER):
    def append_par(table,parname,parvalue):
        if parname not in table:
            table[parname] = [parvalue]
        else:
            table[parname].append(parvalue)
    LOCAL_TABLE_CACHE[table_name] = {'data':{},'header':HITRAN_DEFAULT_HEADER}    
    for item in col.getitems():
        for key in item:
            append_par(LOCAL_TABLE_CACHE[table_name]['data'],key,item[key])

## export HITRAN .par file from a collection
#def export_dotpar(col,filename):
#    with open(filename,'w') as f:
#        for id in col.ids():
#            item = col.__dicthash__[id] # not compatible with future versions
#            line = dump_dotpar(item)
#            f.write(line+'\n')
                
def create_from_buffer(colname,buffer,rstrip=True,lstrip=True,cast=lambda val:val,comment=[]):
    """
    Create a collection from the buffer with just one column, containing the lines from buffer.
    Usually this function is useful for the line-by-line text processing.
    """
    items = []
    for line in buffer.split('\n'):
        if not line.strip(): continue
        if line.lstrip()[0] in comment: continue
        val = line
        if rstrip: val = val.rstrip()
        if lstrip: val = val.lstrip()
        val = cast(val)
        items.append({colname:val})
    col = Collection()
    col.update(items)
    return col
    
def create_from_buffer_multicol(buffer,cast={},duck=True,header=True,comment=[]):
    """
    Create a multicolumn collection from the buffer, using variable-length space delimiter.
    Colnames must obey the same rule as as an ordinary data line.
    Usually this function is useful for the line-by-line text processing.
    """
    
    # Skip first new lines.
    lines = buffer.split('\n')
    for istart,line in enumerate(lines):
        if line.strip(): break
            
    # Get column names.
    if header:
        names = lines[istart].split()        
        istart = istart+1
    else:
        names = ['c%d'%i for i in range(len(lines[0].split()))]        
    
    # Start parsing values.
    items = []    
    for line in lines[istart:]:
        if not line.strip(): continue
        if line.lstrip()[0] in comment: continue
        vals = line.split()  
        # Typing.
        item = {}
        for val,name in zip(vals,names):       
            if name in cast:  # cast typing
                val = cast[name](val)
            elif duck: # duck typing
                try:
                    val = int(val)
                except ValueError as e:
                    try:
                        val = float(val)
                    except ValueError as e:
                        pass
            else: # do nothing, treat as raw string
                pass
            item[name] = val
        items.append(item)
    
    # Create collection from the list of items.
    col = Collection()
    col.update(items)
    col.order = names
    
    return col
  
def join_index(idx1,idx2,inner=True):
    """ 
    
    ALTERNATIVE VERSION, WITH GROUP INDEXES INSTEAD OF COLLECTIONS
    
    Join two collections by key and additional conditions.
    
    idx1, idx2:
        group indexes of col1 and col2.
            
    Inner: 
        Flag to perform the inner join. 
        If False, outer join is performed.
        
    Output:
        Join index which is a list of ID pars (id1,id2) with
        id1 from col1 and id2 from col2.
    """    
                    
    def unroll(ids1,ids2): # unroll to list of pairs
        pairs = []
        for id1 in ids1:
            for id2 in ids2:
                pairs.append((id1,id2))
        return pairs
        
    # Calculate sets of indexvalues.
    keys1 = set(idx1.keys())
    keys2 = set(idx2.keys())
    
    # Prepare intersection for the inner part of join.
    keys_union = keys1.intersection(keys2)
    
    # Start forming the join index.
    idx = ft.reduce(lambda x,y:x+y,
        [unroll(idx1[k],idx2[k]) for k in keys_union])
    
    # If performing outer join, add symmetric difference set to it.
    if not inner:
        keys_diff_12 = keys1-keys2
        keys_diff_21 = keys2-keys1
        
        if keys_diff_12:
            idx += ft.reduce(lambda x,y:x+y,
                [unroll(idx1[k],[None]) for k in keys_diff_12])

        if keys_diff_21:
            idx += ft.reduce(lambda x,y:x+y,
                [unroll([None],idx2[k]) for k in keys_diff_21])
            
    return idx
  
def create_join_index(col1,col2,key,key2=None,inner=True):
    """ 
    Join two collections by key and additional conditions.
    
    col1,col2:
        collections to join together.
            
    key,key2:
        Key is a lambda function that must return comparable 
        object (e.g. tuple or scalar).  
    
    Inner: 
        Flag to perform the inner join. 
        If False, outer join is performed.
        
    Output:
        Join index which is a list of ID pars (id1,id2) with
        id1 from col1 and id2 from col2.
    """    
            
    def prepare_key(key): # convert key to lambda form
        if type(key) is str:
            key = eval('lambda v: v["%s"]'%key)
        elif type(key) in [tuple,list]:
            key = eval('lambda v: (%s)'%(','.join(['v["%s"]'%k for k in key])))
        elif type(key) is type(lambda:None):
            pass
        else:
            raise Exception('unknown key format: %s'%str(key))
        return key
        
    # Prepare keys and convert them to the lambda function format.
    key = prepare_key(key)
    if key2:
        key2 = prepare_key(key2)
    else:
        key2 = key
        
    # Calculate indexes for both collections.
    idx1 = col1.group(key)
    idx2 = col2.group(key2)
    
    idx = join_index(idx1,idx2,inner=inner)
    
    return idx

def join(col1,col2,idx,colnames1=lambda c:c,colnames2=lambda c:'_%s'%c):
    """
    Join two collection based on the result of the join_index/create_join_index functions.
    Colnames 1 and 2 are lambda functions renaming colnames, having the following format:
        colnames = lambda colname: <any name>
    """
    
    col = Collection()
    
    col.order = [colnames1(c) for c in col1.order] + \
                [colnames2(c) for c in col2.order]
    
    for id1,id2 in idx:
        item = {}
        if id1 is not None:
            item1 = col1.getitem(id1)
            for c in item1:
                c_ = colnames1(c)
                item[c_] = item1[c]
        if id2 is not None:
            item2 = col2.getitem(id2)
            for c in item2:
                c_ = colnames2(c)
                if c_ in item:
                    raise Exception('column conflict at join: %s'%c_)
                else:
                    item[c_] = item2[c]
        col.update(item)
    
    return col

def join_index_(*gidxs,inner=True):
    """ multi-collection version of join_index
    
    Join many collections by key and additional conditions.
    
    gidxs:
        group indexes (dicts)
            
    Inner: 
        Flag to perform the inner join. 
        If False, outer join is performed.
        
    Output:
        Join index which is a list of ID pars (id1,id2, ...idn) 
    """    
    
    def get_next(it): # get next value of an iterator
        return next(it,None)
        
    def get_ids(keys,gidxs): # get collection ids from a key tuple
        return [idx[key] if key is not None else [None] for key,idx in zip(keys,gidxs)]
    
    def cproduct(lists): # cartesian product of lists
        return it.product(*lists)
    
    def check_inner(flags):
        return all(flags)

    def check_outer(flags):
        return any(flags)
    
    def iterate(iters,check): # keys must not be None!!!
        curkeys = [get_next(it) for it in iters]
        keyss = []
        while True:
            #print('curkeys>>>',curkeys)
            if all([k is None for k in curkeys]):
                break
            minkey = min([k for k in curkeys if k is not None])
            flags = [k==minkey for k in curkeys]
            if check(flags):
                keys = [minkey if f else None for f in flags]
                keyss.append([minkey,keys])
            curkeys = [get_next(it) if f else k for k,f,it in zip(curkeys,flags,iters)]
        #print('keyss>>>',keyss)
        return keyss
    
    iterate_inner = lambda key_iterators: iterate(key_iterators,check=check_inner)
    iterate_outer = lambda key_iterators: iterate(key_iterators,check=check_outer)
            
    key_iterators = [iter(sorted(idx.keys())) for idx in gidxs]
    
    if inner:
        meta_iterator = iterate_inner(key_iterators)
    else:
        meta_iterator = iterate_outer(key_iterators)
    
    jidx = []
    for val,keys in meta_iterator:
        ids = get_ids(keys,gidxs)
        jidx_ = cproduct(ids)
        jidx += [[val,i] for i in jidx_]
            
    return jidx

#def create_join_index_(cols,keys,inner=True):
#    """ multi-collection version of create_join_index """
#    idxs = [col.group(key) for col,key in zip(cols,keys)]
#    idx = join_index_(idxs,inner=inner)
#    return idx

def join_(jkey,jidx,*cols):
    """ 
    Multi-collection version of join.
    Inputs: 
        jkey -> name of join key (i.e. where to store keyvals of jidx)
                can also be a list/tuple of names with the same length
                as keys of jidx.
        jidx -> join index produced by join_index_()
        cols -> columns to join.
    To reduce/rename column keys now one must use
    slice and map collection methods respectively.
    """

    col_join = Collection()
    
    col_join.order = ft.reduce(lambda x,y: x+y,[c.order for c in cols])
    
    if type(jkey) is str:
        init_item = lambda jval: {jkey:jval}
        col_join.order = [jkey] + col_join.order
    elif type(jkey) in {tuple,list}:
        init_item = lambda jval: {k:v for k,v in zip(jkey,jval)}
        col_join.order = list(jkey) + col_join.order
    else:
        raise Exception('jkey must be either string, list or tuple')
            
    for jval,keys in jidx:        
        item = init_item(jval)
        for ID,col in zip(keys,cols):
            if ID is None:
                continue
            item_ = col.getitem(ID)
            for k in item_:
                if k in item:
                    raise Exception('column conflict at join: %s'%k)
                else:
                    item[k] = item_[k]
        col_join.update(item)
    
    return col_join
    
######################################################################
# RECURSIVE COMPARISON OF THE ARBITRARY DATA STRUCTURES, "DIFF". #####
# SHOULD CORRECTLY COMPARE NESTED STRUCTURES OF DICTS ################
######################################################################

def diff(D1,D2):
    """
    Recursvely compare two nested structures consisting of dict
    objects. These objects should contain elements which are comparable,
    i.e. supporting the "==" operation.
    The result of the diff is ??????
    """
    raise NotImplementedError
    # get keys of data structures
    keys1 = data1.keys()
    keys2 = data2.keys()  
      
    # get three different areas of comparison:
    #  a) D1 minus D2 
    #  b) D2 minus D1
    #  c) D1 intersect D2
    
    output = {}
    
    D1_minus_D2 = set(D1)-set(D2) # always part of output
    if D1_minus_D2: output['left'] = {D1[key] for key in D1_minus_D2}
    
    D2_minus_D1 = set(D2)-set(D1) # always part of output
    if D2_minus_D1: output['right'] = {D2[key] for key in D2_minus_D1}
    
    D1_intersect_D2 = set(D1).intersect(D2) # part of output if there are 
    center = {}
    for key in D1_intersect_D2:
        e1 = D1[key]
        e2 = D2[key]
        if type(e1)==dict and type(e2)==dict:
            out = diff(e1,e2)
            if out: center[key] = out
        elif type(e1)==dict and type(e2)!=dict:
            out = {}

######################################################################
# STORAGE BACKENDS PART ##############################################
# EACH BACKEND IS IMPLEMENTED AS A CONNECTION OBJECT #################
# CONNECTIONS ALLOW READING AND WRITING TO VARIOUS FORMATS: ##########
#       -> DATABASES                                        ##########
#       -> BINARY STORAGES (E.G. HDF5)                      ##########
######################################################################

class StorageConnection:
    
    def __init__(self,*args,**kwargs):
        
        # Connect to the storage.
        self.connect(*args,**kwargs)
        
        # Do post-init business.
        self.__post_init__(*args,**kwargs)
        
    def __post_init__(self,*args,**kwargs):
        pass
    
    def connect(self,*args,**kwargs): # interface
        """
        Open the storage and save the connector.
        """
        raise NotImplementedError
        
    def close(self): # interface
        """
        Close the storage.
        """
        raise NotImplementedError
        
    def get_type_header(self,sql): # interface
        """
        Get the type specification header for the table.
        This specification should be a dictionary,
        where a key is the attribute/column name,
        and value is a type signature supported by Jeanny3.
        """
        raise NotImplementedError
    
    def table_exists(self,tablename): # interface
        """
        Check if the table exists in the storage.
        """
        raise NotImplementedError
        
    def create_table(self,tablename,type_header): # interface
        """
        Create table in the storage with respect to type header.
        Each storage backend can have different name convertion,
        i.e. for RDMBSs the backend expecting to recieve just a 
        name of the table, but for hierarchical storages such as HDF5,
        name can be a sequence of nested "datasets",
        e.g. "dataset1/dataset2/tablename".
        If the storage has only one table, the name can be ignored
        by the backend.
        """
        raise NotImplementedError
    
    def insert_(self,tablename,datamatrix,keynames): # interface
        """
        Insert data matrix to the table.
        Data matrix is a list of rows, each row is 
        either a list, or a tuple having the same order
        and having the same length as keynames.
        """
        raise NotImplementedError
        
    def select_(self,sql): # interface
        """ 
        Must return iterable queryset object.
        Each row of the queryset should be a list or tuple.
        """
        raise NotImplementedError
    
    def insert(self,tablename,col):
        """
        Insert collection into table.
        If type header is not specified in the collection,
        it will be guessed from the data itself.
        To speed up the process, specify the type header manually.
        """
        
        # Get the column schema from collection.
        if col.types is not None:
            # Use existing type header for appending to the database.
            type_header_col = col.types
        else:
            # Find the type header by calling the Collection's method.
            type_header_col = col.get_types()
        
        # Check if the table exists.
        if self.table_exists(tablename):
            # Get the table schema from storage.
            type_header_stor = self.get_type_header('select * from %s'%tablename)
            if len(type_header_stor)>0: # query type header is not empty
                # Get the intersection of keys from two schemas.
                type_header = {
                    key:type_header_col[key] for key in type_header_stor \
                        if key in type_header_col
                }
            else: # query type header is empty
                type_header = type_header_col
        else:
            # Use only the collection type header.
            type_header = type_header_col
            # Create table in the storage.
            self.create_table(tablename,type_header)
        
        assert len(type_header)>0, "no columns to insert"
        
        # Create a datamatrix with fields from type_header.
        keynames = list(type_header)
        datamatrix = col.getrows(keynames)
        
        # Run backend-specific lower-level incert.
        self.insert_(tablename,datamatrix,keynames)
            
    def select(self,sql,chunksize):
        """
        Iteratively select from storage to a sequence of collections.  
        
        ATTENITION: if the lower-level select_ uses a context manager
        (i.e. "with ..." block statement), the chunksize must be not less
        than the maximum number of rows in query result. In other case,
        there will be a context "hang up".
        """
                
        # Get the table schema from storage.
        type_header = self.get_type_header(sql)
        key_header = list(type_header)

        # Get the iterable queryset object.
        #queryset = iter( self.select_(sql) )
        queryset = self.select_(sql)

        # Iterate through storage by chunks.
        while True:
            
            chunk = it.islice(queryset,chunksize) # itertools
            chunk = list(chunk)
            nitems = len(chunk)
        
            if nitems==0: break
            
            # Create and fill collection.
            col = Collection()
            for row in chunk:
                item = {key:val for key,val in zip(key_header,row)}
                col.update(item)

            # Assign tabulation order to collection.
            col.order = key_header
            
            # Assign type header to collection.
            col.types = type_header
                        
            yield col
            
    def command(self,sql):
        """
        Run raw SQL command and print an output.
        """
        raise NotImplementedError

class ClickhouseConnection(StorageConnection):

    def connect(self,
        host,
        database,
        username='default',
        password=None,
        port=None):
            
        import clickhouse_connect as cc
                        
        self.__client__ = cc.get_client(
            host=host, 
            port=port, 
            database=database,
            username=username, 
            password=password,
        )
        
    def close(self):
        self.__client__.close()
        
    def __post_init__(self,*args,**kwargs):
        
        self.__MAP_FROM_CLHS__ = { # left -> ClickHouse type, right -> Python type

            'UInt4': int,
            'UInt8': int,
            'UInt16': int,
            'UInt32': int,
            'UInt64': int,
            'UInt128': int,
            'UInt256': int,

            'Int4': int,
            'Int8': int,
            'Int16': int,
            'Int32': int,
            'Int64': int,
            'Int128': int,
            'Int256': int,

            'Float32': float,
            'Float64': float,
            
            'String': str,
            'FixedString': str,
        }
        
        self.__MAP_TO_CLHS__ = { # left -> Python type, right -> ClickHouse type            
            
            int: 'Int64',
            float: 'Float64',
            str: 'String',            
        }
                
    def get_type_header(self,sql): # interface
        
        # Get client and queryset objects.
        client = self.__client__
        sql_nolimit = re.sub('limit\s+\d+;?','',sql)
        queryset = client.query(sql_nolimit+' limit 1') # WILL NOT WORK WITH ALREADY LIMITED QUERIES, REDO!!!
        
        # Get column types and names
        colnames = queryset.column_names
        coltypes = queryset.column_types
        
        # Map column types to Python types
        coltypes = [
            self.__MAP_FROM_CLHS__[t.__class__.__name__] for t in coltypes]
            
        # Get and return the type header.
        type_header = {key:typ for key,typ in zip(colnames,coltypes)}
        
        #assert len(type_header)>0, "empty type header"
        
        return type_header
    
    def table_exists(self,tablename): # interface
        client = self.__client__
        flag = client.query('exists %s'%tablename).result_rows[0][0]
        return bool(flag)
        
    def create_table(self,tablename,type_header): # interface
        
        # Get client object.
        client = self.__client__
        
        # Get Python->ClickHouse type mapper.
        MAP = self.__MAP_TO_CLHS__
        
        # Get names and types of columns.
        colnames = list(type_header)
        coltypes = [type_header[key] for key in colnames]
        
        # Get ClickHouse column types.
        coltypes_ = [MAP[typ] for typ in coltypes]
        
        # Make a template for create_table command
        #'CREATE TABLE test_command (col_1 String, col_2 DateTime) Engine MergeTree ORDER BY tuple()'
        template = 'CREATE TABLE {TABLENAME} ({COLSPEC}) Engine {ENGINE} ORDER BY {ORDERBY}'
        
        # Create a command and run it using the connector.
        command = template.format(
            TABLENAME=tablename,
            COLSPEC=', '.join(['%s %s'%(key,typ) for key,typ in zip(colnames,coltypes_)]),
            ENGINE='MergeTree',
            ORDERBY='tuple()',            
        )
        
        client.command(command)
    
    def insert_(self,tablename,datamatrix,keynames): # interface
        client = self.__client__
        client.insert(table=tablename,data=datamatrix,column_names=keynames)
        
    def select_(self,sql): # interface
        
        client = self.__client__
        
        # set default encoding
        ENCODING = 'utf-8'
        
        # Create decoding function, to convert from binary to Python string.
        decode = lambda row: tuple(
            [val.decode(ENCODING) if type(val) is bytes else val for val in row])
    
        # Generate sequence of rows in a loop.
        with client.query_rows_stream(sql) as f:
            for row in f:
                row = decode(row)
                yield row
                
    def command(self,sql):
        print( self.__client__.command(sql) )

class HDF5Connection(StorageConnection):
    pass

class SQLiteConnection(StorageConnection):
    pass

######################################################################
# SPREADSHEETS FOR PRETTY-PRINTING ###################################
######################################################################

class Spreadsheet:
    """
    Spreadsheet for pretty-printing collections.
    """
    
    def __init__(self,col=None,header=False):

        if col is None:
            self.cells = []

        elif isinstance(col,Collection):

            assert col.order
            if header==True: header = col.order

            if type(header) in {list,tuple}:
                assert len(header)==len(col.order)
                self.cells = [[SpreadsheetCell(key) for key in header]]
            else:
                self.cells = []
                
            for count,item in enumerate(col.getitems()):
                row = []
                for key in col.order:
                    text = str(item[key]) if key in item else None
                    row.append(SpreadsheetCell(text))
                self.cells.append(row)
                
            self.nrows = count+1
            if header is not None:
                self.nrows += 1
            self.ncols = len(col.order)
        else:
            raise Exception('unknown data type (Collection expected)')
            
    def cell(self,i,j):
        return self.cells[i][j]
        
    def copy(self):
        spread = Spreadsheet()
        spread.nrows = self.nrows
        spread.ncols = self.ncols
        for row in self.cells:
            spread.cells.append([e for e in row])
        return spread

    def select(self,ij_from,ij_to,step_i=1,step_j=1):
        
        spread = Spreadsheet()
        
        i_,j_ = ij_from
        i,j = ij_to
        
        di = -step_i if i_>i else step_i
        dj = -step_j if j_>j else step_j
        
        for ic in range(i_,i+1,di):
            row = []
            for jc in range(j_,j+1,dj):
                row.append(self.cell(ic,jc))
            spread.cells.append(row)
        
        spread.nrows = abs(i-i_+1)
        spread.ncols = abs(j-j_+1)
        
        return spread
        
    def __getitem__(self,index):
        
        index_i,index_j = index
        
        if type(index_i) is int and type(index_j) is int:
            return self.cells[index_i][index_j]
        
        def prepare_indexes(index_i,max_i):            
            if type(index_i) is int:
                i_ = index_i
                i = index_i
                step_i = 0                
            elif type(index_i) is slice:
                i_ = index_i.start if index_i.start is not None else 0
                i = index_i.stop-1 if index_i.stop is not None else max_i
                step_i = index_i.step if index_i.step is not None else 1
                    
            return i_,i,step_i
            
        i_,i,step_i = prepare_indexes(index_i,self.nrows-1)
        j_,j,step_j = prepare_indexes(index_j,self.ncols-1)
        
        return self.select((i_,j_),(i,j),step_i,step_j)
        
    def insert(self,ij_from,spread):
        
        # input is a also Spreadsheet
        datamatrix = spread.cells
        i_,j_ = ij_from
        
        nrows_self = self.nrows
        ncols_self = self.ncols
        
        # extend rows
        di = (i_+spread.nrows)-nrows_self
        if di>0: nrows_self += di
        for _ in range(di):
            self.cells.append([SpreadsheetCell() for _ in range(self.ncols)])

        # extend columns
        dj = (j_+spread.ncols)-ncols_self
        if dj>0: ncols_self += dj
        for ic in range(nrows_self):
            self.cells[ic] += [SpreadsheetCell() for _ in range(dj)]
        
        # loop over cells and assign
        for i_spread in range(spread.nrows):
            i_self = i_ + i_spread
            for j_spread in range(spread.ncols):
                j_self = j_ + j_spread
                cell = datamatrix[i_spread][j_spread] 
                self.cells[i_self][j_self] = cell
                
        self.nrows = nrows_self
        self.ncols = ncols_self

    def set_attributes(self,**kwargs):
        for attr_name in kwargs:
            attr_value = kwargs[attr_name]
            for i in range(self.nrows):
                for j in range(self.ncols):
                    self.cell(i,j).set(attr_name,attr_value)

    def __add__(self,spread): 
        """
        Append another spreadsheet or selection to the bottom.
        """
        assert self.ncols==spread.ncols
        spread_output = self.copy()
        spread_output.nrows += spread.nrows
        for row in spread.cells:
            spread_output.cells.append(row)  
        return spread_output
        
    def latex(self,centering=True,caption=False,label=False):
        
        to_str = lambda a: str(a) if a is not None else ''

        def to_latex(val):    
            buf = to_str(val).strip()
            # Check if value is a math expression.
            if len(buf)>0 and buf[0]==buf[-1]=='$':
                return buf
            # escape underscores
            buf = buf.replace('_','\_')
            return buf
        
        TEMPLATE = """\\documentclass{{article}}

\\usepackage[table]{{xcolor}}
\\usepackage{{booktabs}}
\\usepackage{{adjustbox}}
\\usepackage[
    left=2cm,
    right=2cm,
    top=2cm,
    bottom=2cm,
    bindingoffset=0cm
]{{geometry}}

\\begin{{document}}

\\begin{{table}}
{CENTERING}
\\begin{{adjustbox}}{{width=\\textwidth}}
\\begin{{tabular}}{{{COLSPEC}}}
{LINES}
\\end{{tabular}}
\\end{{adjustbox}}
{CAPTION}
{LABEL}
\\end{{table}}

\\end{{document}}
"""
        CENTERING = '\\centering' if centering else ''
        CAPTION = '\\caption{%s}'%caption if type(caption) is str else ''
        LABEL = '\\label{%s}'%label if type(label) is str else ''
        
        COLSPEC = 'l'*self.ncols
        
        ROWS = []
        
        for row in self.cells:
            
            BUFS = []
            
            BORDERS = []
            
            for i,cell in enumerate(row,start=1):
                
                #BUF = '\\verb|%s|'%cell.text if cell.text else ''
                BUF = to_latex(cell.text) if cell.text else ''
                
                # add enclosing commands
                if cell.font_bold: BUF = '\\textbf{%s}'%BUF
                if cell.font_italic: BUF = '\\textit{%s}'%BUF
                #..... all enclosing commands go here
                
                # add preceding commands
                if cell.text_color: BUF = '\\color{%s}%s'%(cell.text_color,BUF)
                if cell.background_color: BUF = '\\cellcolor{%s}%s'%(cell.background_color,BUF)
                
                # borders
                if cell.border_bottom: BORDERS.append(i)
                if cell.border_top: raise NotImplementedError
                if cell.border_left: raise NotImplementedError
                if cell.border_right: raise NotImplementedError
                
                BUFS.append(BUF)
            
            LINE = ' & '.join(BUFS) + ' \\\\'
            
            # set borders
            for nbrd in BORDERS:
                LINE += '\\cmidrule{%d-%d}'%(nbrd,nbrd)
            
            ROWS.append(LINE)
            
        LINES = '\n'.join(ROWS)
            
        LATEX = TEMPLATE.format(
            CENTERING=CENTERING,
            COLSPEC=COLSPEC,
            LINES=LINES,
            CAPTION=CAPTION,
            LABEL=LABEL,
        )
        
        return LATEX
        
    def tabulate(self):
        from tabulate import tabulate as tab
        print(tab(self.cells))
        
    def __repr__(self):
        return 'Spreadsheet(%dx%d)'%(self.nrows,self.ncols)

class SpreadsheetCell:
    
    ATTRIBUTE_DEFAULTS = {
        'border_top': False,
        'border_bottom': False,
        'border_left': False,
        'border_right': False,
        'font_family': 'Arial',
        'font_size': None, # small, tiny etc.
        'font_bold': False,
        'font_italic': False,
        'text_color': None, # red, black, blue etc...
        'background_color': None, # red, black, blue etc...
    }
    
    def __init__(self,text=None,**kwargs):
        self.text = text
        self.spreadsheed = kwargs.get('spreadsheed',None)
        for attr_name in self.ATTRIBUTE_DEFAULTS:
            setattr(self,
                attr_name,
                kwargs.get(
                    attr_name,
                    self.ATTRIBUTE_DEFAULTS[attr_name]
                )
            )
        
    def set(self,attr_name,attr_value):
        setattr(self,attr_name,attr_value)

    def latex(self):
        return self.text # no fancy stuff so far
        
    def get_attributes(self):
        return {attr_name:getattr(self,attr_name) \
            for attr_name in self.ATTRIBUTE_DEFAULTS}
        
    def __repr__(self):
        return self.text if self.text else ''

######################################################################
# OBJECTS FOR PLOTTING COLLECTIONS ###################################
######################################################################

# USEFUL LINKS ON PLOTTING
# https://matplotlib.org/stable/gallery/subplots_axes_and_figures/subfigures.html
# https://stackoverflow.com/questions/21001088/how-to-add-different-graphs-as-an-inset-in-another-python-graph
# https://stackoverflow.com/questions/16150819/common-xlabel-ylabel-for-matplotlib-subplots
# https://stackoverflow.com/questions/73331245/how-to-adjust-gridspec-spacing
# https://matplotlib.org/stable/api/_as_gen/matplotlib.gridspec.GridSpec.html#matplotlib.gridspec.GridSpec
# https://matplotlib.org/stable/gallery/subplots_axes_and_figures/figure_size_units.html
# https://stackoverflow.com/questions/11637929/remove-padding-from-matplotlib-plotting
# https://stackoverflow.com/questions/3130072/matplotlib-savefig-image-trim
# https://stackoverflow.com/questions/18619880/adjust-figure-margin
# https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.figure.html
# https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.margins.html
# https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.suptitle.html
# https://stackoverflow.com/questions/30108923/whats-the-difference-between-title-and-suptitle
# https://matplotlib.org/3.2.1/api/_as_gen/matplotlib.pyplot.figure.html
# https://github.com/microsoft/vscode-jupyter/issues/9486
# https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.savefig.html
# https://matplotlib.org/stable/users/explain/figure/backends.html#the-builtin-backends
# https://stackoverflow.com/questions/11837979/removing-white-space-around-a-saved-image
# https://blog.rtwilson.com/easily-hiding-items-from-the-legend-in-matplotlib/
# https://github.com/matplotlib/matplotlib/issues/17139/    BUG WITH THE ORDER OF PLOTS IN LEGEND WHEN USING ERRORBAR

###################################
# ABSTRACTIONS FOR THE DATA POINTS
###################################

class PlotData: 
    """
    Abstract class for line-series, supplies data for 2D and 3D plots.
    """
    
    __args__ = []
    
    def __init__(self,*args,**kwargs):
        raise NotImplementedError
        
    def clone(self,**kwargs):
        kwargs_obj = {k:getattr(self,k) for k in self.__args__}
        for k in kwargs:
            kwargs_obj[k] = kwargs[k]
        obj = self.__class__(**kwargs_obj)
        return obj
        
class DataPoints2D(PlotData):
    """
    Points for two-dimensional plots (scatters, line plots, bar plots etc...)
    """
    
    __args__ = ['x','y','s']
    
    def __init__(self,x,y,s=None):
        self.x = x
        self.y = y
        self.s = s # sizes
        
class DataPoints3D(PlotData):
    """
    Points for three-dimensional plots (meshes and surfaces)
    """
    
    __args__ = ['x','y','z','s']
    
    def __init__(self,x,y,z,s=None):
        self.x = x
        self.y = y
        self.z = z
        self.s = s # sizes

class DataText(PlotData):
    """
    Text data for textboxes.
    """
    
    __args__ = ['x','y','text']
    
    def __init__(self,x,y,text):
        
        if type(x) in {float,int}: x = [x]
        if type(y) in {float,int}: y = [y]
        if type(text) in {str}: text = [text]
                    
        assert type(x) in {list,tuple}
        assert type(y) in {list,tuple}
        assert type(text) in {list,tuple}
        
        self.text = text
        self.x = x
        self.y = y

class DataErrorbars(PlotData):
    """
    Text data for textboxes.
    """
    
    __args__ = ['x','y','yerr','xerr']
    
    def __init__(self,x,y,yerr,xerr=None):                            
        self.x = x
        self.y = y
        self.yerr = yerr
        self.xerr = xerr

class DataFillBetween(PlotData):
    """
    Text data for filled rectangle area.
    """
    
    __args__ = ['xmin','xmax','ymin','ymax']
    
    def __init__(self,xmin,xmax,ymin,ymax):                            

        if type(xmin) in {float,int}: xmin = [xmin]
        if type(xmax) in {float,int}: xmax = [xmax]
        if type(ymin) in {float,int}: ymin = [ymin]
        if type(ymax) in {float,int}: ymax = [ymax]
                    
        assert type(xmin) in {list,tuple}
        assert type(xmax) in {list,tuple}
        assert type(ymin) in {list,tuple}
        assert type(ymax) in {list,tuple}
        
        self.xmin = xmin
        self.xmax = xmax
        self.ymin = ymin
        self.ymax = ymax
        
#######################################
# ABSTRACTIONS FOR THE OPTIONS OBJECTS
#######################################

class Options:
    """
    Abstract class for options for layers, axes etc...
    """

    __defaults__ = {}
    
    def __init__(self,**kwargs):
        self.options = self.__defaults__.copy()
        for key in kwargs:
            if key in self.__defaults__:
                self.options[key] = kwargs[key]
            else:
                raise Exception('unknown option "%s"'%key)
            
    def __repr__(self):
        return '%s style class with options:\n%s'%(
            self.__class__.__name__,
            json.dumps(self.options,indent=3)
        )
    
    def __getitem__(self,name):
        return self.options[name]
    
    def __setitem__(self,name,value):
        assert name in self.options
        self.options[name] = value
        
    def __iter__(self):
        return iter(self.options)
    
    def copy(self):
        opts = self.__class__()
        opts.options = self.options.copy()
        return opts

class LineSeriesOptions(Options):
    """
    LineSeries options
    """
    
    __defaults__ = {
        'marker_color': None,
        'marker_style': '.',
        'marker_size': None,
        'marker_alpha': None,
        'line_style': '-',
        'line_color': None,
        'line_width': None,
    }

class TextOptions(Options):
    """
    Text options
    """
    
    __defaults__ = {
        'clip_on': True,
        #'transform': None, # transaxes=>axes coords, default=>data coords
        'floating': False, # # transform=="transaxes"=>axes coords, default=>data coords
        'box_color': None,
        'box_alpha': None,
        'box_style': None, # round, 
        'font_color': None,
        'font_size': None,
        'font_style': None,
        'font_family': None,
    }
    
class ScatterOptions(Options):
    """
    Scatter options
    """
    
    __defaults__ = {
        'marker_color': None,
        'marker_style': '.',
        'marker_size': None,
        'marker_alpha': None,
        'line_style': '-',
        'line_color': None,
        'line_width': None,
    }

class ErrorbarsOptions(Options):
    """
    Errorbars options
    """
    
    __defaults__ = {
        'marker_color': None,
        'marker_style': '.',
        'marker_size': None,
        'marker_alpha': None,
        'line_style': '-',
        'line_color': None,
        'line_width': None,
        'errorbar_color': None,
        'errorbar_width': None,
    }

class FillBetweenOptions(Options):
    """
    Fill options
    """
    
    __defaults__ = {
        'color': None,
        'alpha': None,
    }
    
class AxesOptions(Options):
    """
    Axes options
    """

    __defaults__ = {
        'title': None,
        'title_pad': None,
        'title_location': None,
        'title_font_family': None,
        'title_font_size': None,
        'title_font_style': None,
        'legend_on': True,
        'legend_font_size': None,
        'legend_font_style': None,
        'legend_font_family': None,
        'legend_markerscale': None,
        'legend_borderpad': None,
        'legend_location': None,
        'grid_on': True,
        'grid_which': None, # major {'major', 'minor', 'both'}
        'grid_axis': None, # both {'both', 'x', 'y'}
        'x_axis_limits': None,
        'y_axis_limits': None,
        'z_axis_limits': None,
        'x_axis_label': None,
        'y_axis_label': None,
        'z_axis_label': None,
        'x_axis_label_show': True,
        'y_axis_label_show': True,
        'z_axis_label_show': True,
        'x_axis_label_font_family': None,
        'y_axis_label_font_family': None,
        'z_axis_label_font_family': None,
        'x_axis_label_font_size': None,
        'y_axis_label_font_size': None,
        'z_axis_label_font_size': None,
        'x_axis_label_font_style': None,
        'y_axis_label_font_style': None,
        'z_axis_label_font_style': None,
        'x_axis_logscale_on': None,
        'y_axis_logscale_on': None,
        'z_axis_logscale_on': None,
        'x_axis_ticks_labels': None,
        'y_axis_ticks_labels': None,
        'z_axis_ticks_labels': None,
        'x_axis_ticks_font_family': None,
        'y_axis_ticks_font_family': None,
        'z_axis_ticks_font_family': None,
        'x_axis_ticks_font_size': None,
        'y_axis_ticks_font_size': None,
        'z_axis_ticks_font_size': None,
    }

class FigureOptions(Options):
    """
    Figure options
    """
    
    __defaults__ = {
        'size': None,
        'dpi': None,
        'suptitle': None,
        'suptitle_x': None, # def. 0.5
        'suptitle_y': None, # def. 0.98
        'suptitle_font_family': None,
        'suptitle_font_size': None,
        'suptitle_font_style': None,
        'margins': None,
        'margins_x': None,
        'margins_y': None,
        'margins_tight': True,
        'xlabel': None,
        'xlabel_font_family': None,
        'xlabel_font_size': None,
        'xlabel_font_style': None,
        'xlabel_x': 0.5,
        'xlabel_y': 0.04,
        'ylabel': None,
        'ylabel_font_family': None,
        'ylabel_font_size': None,
        'ylabel_font_style': None,
        'ylabel_x': 0.04,
        'ylabel_y': 0.5,
    }

class GridSpecOptions(Options):
    """
    GridSpec options
    """
    
    __defaults__ = {
        'space_between_plots_height': None,
        'space_between_plots_width': None,
    }
    
########################################
# ABSTRACTIONS FOR THE PLOTTING ROUTINES
########################################

class Layer:
    """
    Abstract class for the axes objects (line plots, scatters, ...).
    """
    
    __in_legend__ = True

    __options_class__ = Options
        
    def __init__(self,data,name=None,options=None):
        
        if options is None: options = self.__options_class__()
        
        assert isinstance(options,self.__options_class__)
        
        self.assert_data(data)
        self.data = data
        self.name = name     
        self.options = options
        
    def clone(self,**kwargs):
        
        obj = self.__class__(
            data = self.data,
            name = self.name
        )
                
        obj.options = self.options.copy()
        for key in kwargs:
            obj.options[key] = kwargs[key]
        
        return obj
        
    def assert_data(self):
        raise NotImplementedError
        
    def plot(self,ax): # main drawing function, like plot/scatter/surface etc...
        raise NotImplementedError
        
class LineSeries(Layer):
    """
    Simple two-dimentional line or point plot.
    """
    
    __options_class__ = LineSeriesOptions
    
    def assert_data(self,data):
        assert isinstance(data,DataPoints2D)
        
    def plot(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
        
        x = self.data.x
        y = self.data.y
                
        MAP = {
            'marker_color': 'markerfacecolor',
            'marker_style': 'marker',
            'marker_size': 'markersize',
            'marker_alpha': 'alpha',
            'line_style': 'linestyle',
            'line_color': 'color',
            'line_width': 'linewidth',
        }
        
        opts = self.options
        opts = {MAP[key]:opts[key] \
                for key in opts if opts[key] is not None}
        
        if DEBUG: print('LineSeries.plot:',opts)

        mpl_ax.plot(x,y,**opts)

class Scatter(Layer):
    """
    Simple two-dimentional line or point plot.
    """
    
    __options_class__ = ScatterOptions # stub, not implemented yet
    
    def assert_data(self,pnts):
        assert isinstance(pnts,DataPoints2D)

    def plot(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
        
        x = self.data.x
        y = self.data.y
        s = self.data.s
                
        MAP = {
            'marker_color': 'markercolor',
            'marker_style': 'marker',
            'marker_size': 'markersize',
            'marker_alpha': 'alpha',
            'line_style': 'linestyle',
            'line_color': 'color',
            'line_width': 'linewidth',
        }
        
        opts = self.options
        opts = {MAP[key]:opts[key] \
                for key in opts if opts[key] is not None}
        
        if s is not None: opts['s'] = s
        
        if DEBUG: print('Scatter.plot:',opts)
            
        mpl_ax.scatter(x,y,**opts)

class Errorbars(Layer):
    """
    Simple two-dimentional line or point plot.
    """
    
    __options_class__ = ErrorbarsOptions
    
    def assert_data(self,data):
        assert isinstance(data,DataErrorbars)
        
    def plot(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
        
        x = self.data.x
        y = self.data.y
        yerr = self.data.yerr
        xerr = self.data.xerr
        
        MAP = {
            'marker_color': 'markerfacecolor',
            'marker_style': 'marker',
            'marker_size': 'markersize',
            'marker_alpha': 'alpha',
            'line_style': 'linestyle',
            'line_color': 'color',
            'line_width': 'linewidth',
            'errorbar_color': 'ecolor',
            'errorbar_width': 'elinewidth',
        }
                
        opts = self.options
        kwargs = {MAP[key]:opts[key] \
                for key in opts if opts[key] is not None}
        
        # The following is the dirty hack dealing with the bug of Matplotlib
        # when the errorbars messes up the order of legend markers.
            
        dummy_plot_kwargs = {}
        if opts['marker_style'] is not None: 
            dummy_plot_kwargs['marker'] = opts['marker_style']
        if opts['marker_color'] is not None: 
            dummy_plot_kwargs['color'] = opts['marker_color']
        if opts['marker_size'] is not None: 
            dummy_plot_kwargs['markersize'] = opts['marker_size']
        #if opts['line_style'] is not None: 
        #    dummy_plot_kwargs['linestyle'] = opts['line_style']
        #if opts['line_color'] is not None: 
        #    dummy_plot_kwargs['color'] = opts['line_color']
        if opts['line_width'] is not None: 
            dummy_plot_kwargs['linewidth'] = opts['line_width']

        dummy_plot_kwargs['linestyle'] = '-'
            
        if DEBUG: print('Errorbars.plot dummy:',dummy_plot_kwargs)

        obj, = mpl_ax.plot([],[],**dummy_plot_kwargs)
        if opts['marker_color'] is None: 
            color = obj.get_color()
            kwargs['markerfacecolor'] = color
            kwargs['color'] = color
            
        kwargs['zorder'] = -1
        
        if DEBUG: print('Errorbars.plot:',kwargs)
        
        mpl_ax.errorbar(x,y,yerr,xerr,**kwargs)

class FillBetween(Layer):
    """
    Discrete area filling between multiple X and Y bounds.
    """
    
    __options_class__ = FillBetweenOptions
    
    def assert_data(self,data):
        assert isinstance(data,DataFillBetween)
        
    def plot(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
        
        xmin = self.data.xmin
        xmax = self.data.xmax
        ymin = self.data.ymin
        ymax = self.data.ymax
        
        n = len(xmin)
                
        MAP = {
            'color': 'color',
            'alpha': 'alpha',
        }
        
        opts = self.options
        opts = {MAP[key]:opts[key] \
                for key in opts if opts[key] is not None}
        
        if DEBUG: print('FillBetween.plot:',opts)

        # get delta
        xx = sorted(xmin+xmax)
        delta = min([x2-x1 for x1,x2 in zip(xx[:-1],xx[1:])])
        
        for i,x_,x,y_,y in zip(range(n),xmin,xmax,ymin,ymax):
            if i==0:
                mpl_ax.fill_between([x_,x],y_,y,**opts)
            else:
                mpl_ax.fill_between([x_,x],y_,y,label='_nolegend_',**opts)
        
class Surface(Layer):
    """
    Simple three-dimentional surface.
    """
    
    __options_class__ = type(None) # stub, not implemented yet
    
    def assert_data(self,data):
        assert isinstance(data,DataPoints2D)

    def plot(self,ax):
        raise NotImplementedError

class Text(Layer):
    """
    Text layer tied to the coordinates.
    """
    
    __in_legend__ = False
    
    __options_class__ = TextOptions
    
    def assert_data(self,data):
        assert isinstance(data,DataText)
        
    def plot(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
        
        text = self.data.text
 
        x = self.data.x
        y = self.data.y
        
        opts = self.options

        kwargs = {}

        # Handle font options.
        MAP = {
            'font_color': 'color',
            'font_size': 'size',
            'font_style': 'style',
            'font_family': 'family',
        }
        fontdict = {MAP[key]:opts[key] \
                    for key in MAP if opts[key] is not None} 
        if fontdict:
            kwargs['fontdict'] = fontdict

        # Handle box options.
        MAP = {
            'box_color': 'facecolor',
            'box_alpha': 'alpha',
            'box_style': 'boxstyle',
        }
        props = {MAP[key]:opts[key] \
                    for key in MAP if opts[key] is not None}
        if props:
            kwargs['bbox'] = props
        
        # Handle transform of the x,y coords
        #if opts['transform'] is not None:
        #    if opts['transform'].lower()=='transaxes':
        #        kwargs['transform'] = mpl_ax.transAxes
        #    else:
        #        raise Exception('error in parameter:','transform')
        if opts['floating']:
            kwargs['transform'] = mpl_ax.transAxes
        
        # Handle other options.
        
        MAP = {
            'clip_on': 'clip_on',
        }
        
        opts = self.options
        opts = {MAP[key]:opts[key] \
                for key in MAP if opts[key] is not None}
        kwargs.update(opts)

        if DEBUG: print('Text.plot:',kwargs)
        
        #mpl_ax.text(x,y,text,**kwargs)
        for x_,y_,text_ in zip(x,y,text):
            mpl_ax.text(x_,y_,text_,**kwargs)
        
###################################
# ABSTRACTIONS FOR THE AXES OBJECTS
###################################

class Axes:
    """
    Abstract class for the multidimensional axes.
    """
    
    def __init__(self,layers,options=None):
        
        if options is None: options = AxesOptions()
        
        assert isinstance(options,AxesOptions)
        self.options = options #includes x/y limits and other axes-specific things
        
        self.layers = []
        if isinstance(layers,Layer):
            layers = [layers]
        
        self.named_layers = {}
        
        self.insert_layers(layers)
                    
        self.insets = []
        
    def insert_layers(self,layers):    
        for layer in layers:
            assert isinstance(layer,Layer)
            self.layers.append(layer)
            if layer.name is not None:
                self.named_layers[layer.name] = layer
        
    def add_inset(self,ins):
        assert isinstance(ins,AxesInset)
        self.insets.append(ins)
            
    def clone(self,**kwargs):
        obj = self.__class__([])
        layers = [layer.clone() for layer in self.layers]
        obj.insert_layers(layers)
        obj.options = self.options.copy()
        for key in kwargs:
            obj.options[key] = kwargs[key]
        return obj
        
    def add_layers(self,layers):
        layer_names = [layer.name for layer in layers if layer.name]
        common_names = set(self.named_layers).intersection(layer_names)
        if common_names:
            raise Exception('already have layers with such names:',
                list(common_names))
        for layer in layers:
            self.layers.append(layer)
            if layer.name is not None:
                self.named_layers[layer.name] = layer
                
    def add_layer(self,layer):
        self.add_layers([layer])
            
    def get_layer(self,layer_name):
        return self.named_layers[layer_name]
    
    def __getitem__(self,layer_name):
        return self.get_layer(layer_name)
    
    def plot_(self,mpl_ax):
        
        DEBUG = SETTINGS['DEBUG']
                
        opts = self.options
        
        # Drawing layers.
        for layer in self.layers:
            layer.plot(mpl_ax)
            
        # Handling title.
        if 'title' in opts and opts['title']:
            
            title_kwargs = {}
            title_fontdict = {}
                        
            MAP = {
                'title_pad': 'pad',
                'title_location': 'loc',
            }
            
            for key in MAP:
                if opts[key] is not None:
                    title_kwargs[MAP[key]] = opts[key]
            
            MAP = {
                'title_font_size': 'size',
                'title_font_family': 'family',
                'title_font_style': 'style',
            }
            
            for key in MAP:
                if opts[key] is not None:
                    title_fontdict[MAP[key]] = opts[key]
                    
            if title_fontdict: title_kwargs['fontdict'] = title_fontdict
            
            mpl_ax.set_title(opts['title'],**title_kwargs)

        # Handling legend.
        if 'legend_on' in opts and opts['legend_on']:
            
            leg = [layer.name for layer in self.layers if layer.__in_legend__]
            
            leg_kwargs = {}
            leg_prop = {}
            
            # more legend font params - see link:
            #https://www.freecodecamp.org/news/how-to-change-legend-fontsize-in-matplotlib/
            
            # "PROP"
            
            MAP = {
                'legend_font_size': 'size',
                'legend_font_family': 'family',
                'legend_font_style': 'style',
            }
            
            for key in MAP:
                if opts[key] is not None:
                    leg_prop[MAP[key]] = opts[key]
                            
            if leg_prop: leg_kwargs['prop'] = leg_prop

            MAP = {
                'legend_borderpad': 'borderpad',
                'legend_location': 'loc',
                'legend_markerscale': 'markerscale',
            }

            # OTHER KWARGS
            
            for key in MAP:
                if opts[key] is not None:
                    leg_kwargs[MAP[key]] = opts[key]

            # more legend font params - see link:
            # https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.legend.html
            
            if DEBUG: print('Axes.legend:',leg,leg_kwargs)
            
            mpl_ax.legend(leg,**leg_kwargs)
        
        # Handling grid lines.
        if 'grid_on' in opts and opts['grid_on']:

            grid_kwargs = {'visible':True}
            
            if DEBUG: print('Axes.grid:',grid_kwargs)
            
            mpl_ax.grid(**grid_kwargs)
        
        # Handling axes.
        for axes_name in ['x','y','z']:
            
            # set axes limits
            opt_name = '%s_axis_limits'%axes_name
            if opt_name in opts:
                opt_value = opts[opt_name]
                if opt_value: 
                    if DEBUG: print('Axes.set_%slim:'%axes_name,opt_value)
                    getattr(mpl_ax,'set_%slim'%axes_name)(opt_value)
                    
            # set axes labels
            label = opts['%s_axis_label'%axes_name]
            label_show = opts['%s_axis_label_show'%axes_name]
            if label and label_show:
            
                label = opts['%s_axis_label'%axes_name]
                
                # collect font properties
                label_kwargs = {}
                fontdict = {}
                    
                font_opt_name = '%s_axis_label_font_size'%axes_name
                if font_opt_name in opts: fontdict['size'] = opts[font_opt_name]

                font_opt_name = '%s_axis_label_family'%axes_name
                if font_opt_name in opts: fontdict['family'] = opts[font_opt_name]
                        
                if fontdict: label_kwargs['fontdict'] = fontdict
                    
                if DEBUG: print('Axes.set_%slabel:'%axes_name,label,label_kwargs)
                getattr(mpl_ax,'set_%slabel'%axes_name)(label,**label_kwargs)
                
            # set axes ticks
            tick_kwargs = {}
            font_opt_name = '%s_axis_ticks_font_family'%axes_name
            if opts[font_opt_name] is not None: tick_kwargs['labelfontfamily'] = opts[font_opt_name]
            font_opt_name = '%s_axis_ticks_font_size'%axes_name
            if opts[font_opt_name] is not None: tick_kwargs['labelsize'] = opts[font_opt_name]
            if tick_kwargs: mpl_ax.tick_params(axis=axes_name,**tick_kwargs)

            # set axes scales
            opt_name = '%s_axis_logscale_on'%axes_name
            opt_value = opts[opt_name]
            if opt_value:
                if DEBUG: print('Axes.set_%sscale:'%axes_name,opt_value)
                getattr(mpl_ax,'set_%sscale'%axes_name)('log')
                    
        # Handling inset axes.
        for inset in self.insets:
            inset_ax = inset.axes
            inset_pos = inset.position
            inset_size = inset.size
            inset_mpl_ax = mpl_ax.inset_axes(inset_pos+inset_size)
            inset_ax.plot_(inset_mpl_ax)
            
    def plot(self,**kwargs):
        Figure(self,options=FigureOptions(**kwargs)).plot()

    def savefig(self,fname,dpi=150,pad_inches=0.03,bbox_inches="tight",backend='agg',
                facecolor='white',transparent=False,**kwargs):
        Figure(self,options=FigureOptions(**kwargs)).savefig(
            fname,dpi=dpi,pad_inches=pad_inches,bbox_inches=bbox_inches,
            backend=backend,facecolor=facecolor,transparent=transparent
        )

class AxesInset:
    """
    Class for insets.
    """
    
    def __init__(self,ax,position,size): # position and size are in the relative units
        
        assert isinstance(ax,Axes)
        
        self.axes = ax
        self.position = position
        self.size = size
                    
#######################################
# ABSTRACTIONS FOR THE GRIDSPEC OBJECTS
#######################################

class GridSpec:
    """
    Class for defining the gridspec.
    """
        
    def __init__(self,ni,nj,options=None):
        
        if options is None: options = GridSpecOptions()
        
        assert isinstance(options,GridSpecOptions)
        self.options = options
    
        self.ni = ni
        self.nj = nj
        self.items = []
    
    def __setitem__(self,ij,obj):
        
        def getindexbounds(ispec,ni):
            if type(ispec) is int:
                istart = ispec
                iend = ispec
            elif type(ispec) is slice:
                istart = ispec.start
                iend = ispec.stop if ispec.stop is not None else ni-1
            else:
                raise Exception('wrong index specification (%s)'%str(ispec))
            return istart,iend
        
        i,j = ij
        istart,iend = getindexbounds(i,self.ni)
        jstart,jend = getindexbounds(j,self.nj)
        gsitem = GridSpecItem(obj,istart,iend,jstart,jend)
        self.items.append(gsitem)
        
    def __plot_subgridspec__(self,fig,gs,mpl_gs=None):
        
        DEBUG = SETTINGS['DEBUG']
        
        if DEBUG: print('GridSpec.__plot_subgridspec__:',gs.ni,gs.nj)
        
        def get_gridspec_kwargs(gs):
            opts = gs.options
            kwargs = {}
            if opts['space_between_plots_height']: kwargs['hspace'] = opts['space_between_plots_height']
            if opts['space_between_plots_width']: kwargs['wspace'] = opts['space_between_plots_width']
            return kwargs
        
        if mpl_gs is None:
            import matplotlib.gridspec as gridspec
            mpl_gs_kwargs = get_gridspec_kwargs(gs)
            mpl_gs = gridspec.GridSpec(gs.ni,gs.nj,figure=fig,**mpl_gs_kwargs)

        for item in gs.items:
            
            obj = item.object

            slice_i = slice(item.istart,item.iend+1)
            slice_j = slice(item.jstart,item.jend+1)
            mpl_gs_slice = mpl_gs[slice_i,slice_j]

            if isinstance(obj,Axes):
                mpl_ax = fig.add_subplot(mpl_gs_slice)
                if DEBUG: print(
                    'GridSpec.<plot_axes>:',item.istart,item.iend,item.jstart,item.jend)
                obj.plot_(mpl_ax)
            elif isinstance(obj,GridSpec):
                if DEBUG: print('GridSpec.<plot_subgridspec>')
                sub_gs_kwargs = get_gridspec_kwargs(obj)
                mpl_sub_gs = mpl_gs_slice.subgridspec(obj.ni,obj.nj,**sub_gs_kwargs)
                self.__plot_subgridspec__(fig,obj,mpl_sub_gs)
        
    def plot(self,fig):
        if SETTINGS['DEBUG']: print('GridSpec.plot')
        self.__plot_subgridspec__(fig,self)
        
class GridSpecItem:
    """
    Gridspec item.
    The init function takes i,j specifiers as an input.
    The i,j specifiers can be either integers or slice objects.
    """
    
    def __init__(self,obj,istart,iend,jstart,jend):
        
        assert isinstance(obj,GridSpec) or isinstance(obj,Axes)
        
        self.object_type = type(obj)
        self.object = obj
        self.istart = istart
        self.iend = iend
        self.jstart = jstart
        self.jend = jend

#######################################
# ABSTRACTIONS FOR THE FIGURE OBJECTS
#######################################

class Figure:
    """
    Class for defining the figure.
    """
    
    def __init__(self,gridspec,options=None):
        
        if options is None: options = FigureOptions()
                                
        #assert isinstance(gridspec,GridSpec)
        assert (isinstance(gridspec,GridSpec) or isinstance(gridspec,Axes))
        
        if isinstance(gridspec,Axes):
            axes = gridspec
            gridspec = GridSpec(1,1)
            gridspec[0,0] = axes
        
        self.grispec = gridspec
        self.options = options
        
    def plot_(self):
        
        DEBUG = SETTINGS['DEBUG']
        
        import matplotlib as mpl
        import matplotlib.pyplot as plt

        if DEBUG: print('Figure.plot_')
        
        opts = self.options
        kwargs = {}
        
        if opts['dpi']: kwargs['dpi'] = opts['dpi']
        if opts['size']: kwargs['figsize'] = opts['size']
        
        mpl_fig = plt.figure(**kwargs)
        self.grispec.plot(mpl_fig)
        
        opts = self.options
        
        #plt.tight_layout() # doesn't work?

        if opts['xlabel']: 
            xlabel_x = opts['xlabel_x']
            xlabel_y = opts['xlabel_y']
            kwargs = dict(ha='center')
            fontdict = {}
            if opts['xlabel_font_family'] is not None: fontdict['family'] = opts['xlabel_font_family']
            if opts['xlabel_font_size'] is not None: fontdict['size'] = opts['xlabel_font_size']
            if opts['xlabel_font_style'] is not None: fontdict['style'] = opts['xlabel_font_style']
            if fontdict: kwargs['fontdict'] = fontdict
            mpl_fig.text(xlabel_x,xlabel_y,opts['xlabel'],**kwargs)
        
        if opts['ylabel']: 
            ylabel_x = opts['ylabel_x']
            ylabel_y = opts['ylabel_y']
            kwargs = dict(va='center',rotation='vertical')
            fontdict = {}
            if opts['ylabel_font_family'] is not None: fontdict['family'] = opts['ylabel_font_family']
            if opts['ylabel_font_size'] is not None: fontdict['size'] = opts['ylabel_font_size']
            if opts['ylabel_font_style'] is not None: fontdict['style'] = opts['ylabel_font_style']
            if fontdict: kwargs['fontdict'] = fontdict            
            mpl_fig.text(ylabel_x,ylabel_y,opts['ylabel'],**kwargs)
        
        if opts['suptitle']: 
            kwargs = {}
            if opts['suptitle_x']: kwargs['x'] = opts['suptitle_x']
            if opts['suptitle_y']: kwargs['y'] = opts['suptitle_y']
            fontprops = {}
            if opts['suptitle_font_family']: fontprops['family'] = opts['suptitle_font_family']
            if opts['suptitle_font_size']: fontprops['size'] = opts['suptitle_font_size']
            if opts['suptitle_font_style']: fontprops['style'] = opts['suptitle_font_style']
            if fontprops: kwargs['fontproperties'] = fontprops
            plt.suptitle(opts['suptitle'],**kwargs)
        
        if opts['margins']:
            kwargs = {}
            if opts['margins_x']: kwargs['x'] = opts['margins_x']
            if opts['margins_y']: kwargs['y'] = opts['margins_y']
            if opts['margins_tight'] is not None: kwargs['tight'] = opts['margins_tight']
            mrg = [opts['margins']] if type(opts['margins']) in {float} else opts['margins']
            plt.margins(*mrg,**kwargs)

    def plot(self):
        self.plot_() # first matplotlib imports should be in plot_
        import matplotlib.pyplot as plt
        plt.show()
        
    def savefig(self,fname,
                dpi=150,pad_inches=0.03,bbox_inches="tight",backend='agg',
                facecolor='white',transparent=False):
        self.plot_() # first matplotlib imports should be in plot_
        import matplotlib.pyplot as plt
        plt.savefig(fname=fname,dpi=dpi,pad_inches=pad_inches,
            bbox_inches=bbox_inches,backend=backend,facecolor=facecolor,transparent=transparent)
        plt.close() # don't display plot, just close
